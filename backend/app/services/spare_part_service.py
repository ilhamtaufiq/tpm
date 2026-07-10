import os
import shutil
import uuid
import io
import openpyxl
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Optional, Dict, Any, List

from sqlalchemy import func, or_, case, text
from sqlalchemy.orm import Session
from fastapi import HTTPException, status, UploadFile
from app.config import settings
from app.utils.sparepart_stock import ALWAYS_READY_STOCK, is_always_ready_stock
from app.realtime import publish_realtime_event

from app.models.bengkel import SparePart
from app.schemas.bengkel import SparePartCreate, SparePartUpdate


class SparePartService:
    """Service for spare part inventory management."""

    def __init__(self, db: Session):
        self.db = db

    def _emit_change(self, action: str, spare_part: Optional[SparePart] = None, spare_part_id: Optional[int] = None) -> None:
        entity_id = spare_part.id if spare_part else spare_part_id
        publish_realtime_event(
            event=f"master.spare_parts.{action}",
            scope="master",
            entity="spare_parts",
            action=action,
            entity_id=entity_id,
            data={
                "kode": getattr(spare_part, "kode_part", None) if spare_part else None,
                "nama": getattr(spare_part, "nama", None) if spare_part else None,
            },
        )

    def generate_next_kode(self, offset: int = 0) -> str:
        """Generate unique spare part code.
        
        Format: SPR + YYMM + NNNN (e.g. SPR26040001)
        Prefix is 7 chars: 'SPR' (3) + 'YYMM' (4)
        
        Args:
            offset: Additional offset to add to the counter (used during bulk import
                     to avoid duplicate kodes when newly added records aren't yet 
                     visible via DB queries).
        """
        today = datetime.now()
        prefix = "SPR"
        date_str = today.strftime("%y%m")
        full_prefix = f"{prefix}{date_str}"  # e.g. "SPR2604" = 7 chars

        last = (
            self.db.query(SparePart)
            .filter(SparePart.kode.like(f"{full_prefix}%"))
            .order_by(SparePart.kode.desc())
            .first()
        )

        if last:
            # Extract numeric suffix AFTER the prefix (not just last 4 chars)
            num_str = last.kode[len(full_prefix):]  # e.g. "0001" or "10000"
            try:
                last_num = int(num_str)
            except (ValueError, IndexError):
                last_num = 0
            new_num = last_num + 1 + offset
        else:
            new_num = 1 + offset

        return f"{full_prefix}{new_num:04d}"

    def create(self, data: SparePartCreate, user_id: Optional[int] = None) -> SparePart:
        """Create a new spare part."""
        # Check duplicate name
        existing = (
            self.db.query(SparePart)
            .filter(
                SparePart.nama == data.nama,
                SparePart.deleted_at.is_(None),
            )
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Spare part dengan nama '{data.nama}' sudah ada",
            )

        # Generate kode if not provided
        kode = data.kode if data.kode else self.generate_next_kode()

        # Check duplicate kode
        existing_kode = self.db.query(SparePart).filter(SparePart.kode == kode).first()
        if existing_kode:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Kode spare part '{kode}' sudah digunakan",
            )

        spare_part = SparePart(
            kode=kode,
            nama=data.nama,
            kode_part=data.kode_part,
            kode_ean=data.kode_ean,
            kategori=data.kategori,
            merek=data.merek,
            satuan=data.satuan,
            stok=data.stok,
            stok_minimum=data.stok_minimum,
            harga_beli=data.harga_beli,
            harga_jual=data.harga_jual,
            lokasi_rak=data.lokasi_rak,
            catatan=data.catatan,
            gambar=data.gambar,
        )

        self.db.add(spare_part)
        self.db.commit()
        self.db.refresh(spare_part)
        self._emit_change("created", spare_part)

        return spare_part

    def get_by_id(self, spare_part_id: int) -> SparePart:
        """Get spare part by ID."""
        spare_part = (
            self.db.query(SparePart)
            .filter(
                SparePart.id == spare_part_id,
                SparePart.deleted_at.is_(None),
            )
            .first()
        )
        if not spare_part:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Spare part tidak ditemukan",
            )
        return spare_part

    def get_by_kode(self, kode: str) -> Optional[SparePart]:
        """Get spare part by kode."""
        return (
            self.db.query(SparePart)
            .filter(
                SparePart.kode == kode,
                SparePart.deleted_at.is_(None),
            )
            .first()
        )

    def update(self, spare_part_id: int, data: SparePartUpdate) -> SparePart:
        """Update spare part정보."""
        spare_part = self.get_by_id(spare_part_id)

        update_data = data.model_dump(exclude_unset=True)
        # Check duplicate name if changing
        if "nama" in update_data and update_data["nama"] != spare_part.nama:
            existing = (
                self.db.query(SparePart)
                .filter(
                    SparePart.nama == update_data["nama"],
                    SparePart.id != spare_part_id,
                    SparePart.deleted_at.is_(None),
                )
                .first()
            )
            if existing:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Spare part dengan nama '{update_data['nama']}' sudah ada",
                )

        for key, value in update_data.items():
            if hasattr(spare_part, key):
                setattr(spare_part, key, value)

        self.db.commit()
        self.db.refresh(spare_part)
        self._emit_change("updated", spare_part)

        return spare_part

    def upload_image(self, spare_part_id: int, file: UploadFile) -> SparePart:
        """Upload image for a spare part."""
        spare_part = self.get_by_id(spare_part_id)
        
        # Ensure upload directory exists
        resolved_path = os.path.realpath(settings.upload_full_path)
        img_dir = os.path.join(resolved_path, "spare_parts")
        os.makedirs(img_dir, exist_ok=True)
        
        # Generate unique filename
        ext = os.path.splitext(file.filename)[1].lower()
        file_id = str(uuid.uuid4())
        new_filename = f"{file_id}{ext}"
        
        # Relative path for DB
        file_path = f"spare_parts/{new_filename}"
        
        # Absolute path for filesystem
        full_path = os.path.join(img_dir, new_filename)
        
        # Delete old image if exists
        if spare_part.gambar:
            old_path = os.path.join(resolved_path, spare_part.gambar.replace("/", os.sep))
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except:
                    pass
        
        # Save file
        with open(full_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Update DB
        spare_part.gambar = file_path
        self.db.commit()
        self.db.refresh(spare_part)
        self._emit_change("image_uploaded", spare_part)
        
        return spare_part

    def get_list(
        self,
        skip: int = 0,
        limit: int = 20,
        search: Optional[str] = None,
        kategori: Optional[str] = None,
        merek: Optional[str] = None,
        low_stock_only: bool = False,
        sort_by: str = "nama",
        sort_order: str = "asc",
    ) -> Dict[str, Any]:
        """Get list of spare parts with pagination and filters."""
        query = self.db.query(SparePart).filter(SparePart.deleted_at.is_(None))

        # Search filter
        if search:
            search_filter = f"%{search}%"
            query = query.filter(
                or_(
                    SparePart.nama.ilike(search_filter),
                    SparePart.kode.ilike(search_filter),
                    SparePart.kode_part.ilike(search_filter),
                    SparePart.kode_ean.ilike(search_filter),
                    SparePart.merek.ilike(search_filter),
                )
            )

        # Category filter
        if kategori:
            query = query.filter(SparePart.kategori == kategori)

        # Brand filter
        if merek:
            query = query.filter(SparePart.merek == merek)

        # Low stock filter
        if low_stock_only:
            query = query.filter(
                SparePart.stok <= SparePart.stok_minimum,
                SparePart.stok != ALWAYS_READY_STOCK
            )

        # Count total
        total = query.count()

        # Sorting
        if sort_by == "penjualan":
            from app.models.bengkel import DetailTransaksiSpareParts
            query = (
                query.outerjoin(DetailTransaksiSpareParts)
                .group_by(SparePart.id)
            )
            if sort_order == "desc":
                query = query.order_by(func.sum(DetailTransaksiSpareParts.qty).desc())
            else:
                query = query.order_by(func.sum(DetailTransaksiSpareParts.qty).asc())
        else:
            sort_column = getattr(SparePart, sort_by, SparePart.nama)
            if sort_order == "desc":
                query = query.order_by(sort_column.desc())
            else:
                query = query.order_by(sort_column.asc())

        # Pagination
        spare_parts = query.offset(skip).limit(limit).all()

        # Calculate pages
        pages = (total + limit - 1) // limit if limit > 0 else 1

        return {
            "data": spare_parts,
            "total": total,
            "page": (skip // limit) + 1 if limit > 0 else 1,
            "size": limit,
            "pages": pages,
        }


    def delete(self, spare_part_id: int) -> bool:
        """Soft delete spare part."""
        spare_part = self.get_by_id(spare_part_id)

        # Check if has stock
        if spare_part.stok > 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Tidak dapat menghapus spare part yang masih memiliki stok",
            )

        spare_part.deleted_at = datetime.now()
        self.db.commit()
        self._emit_change("deleted", spare_part)

        return True

    def update_stock(
        self,
        spare_part_id: int,
        quantity: int,
        operation: str = "add",
    ) -> SparePart:
        """Update spare part stock.

        Args:
            spare_part_id: ID of spare part
            quantity: Amount to add/subtract
            operation: 'add' or 'subtract'
        """
        spare_part = self.get_by_id(spare_part_id)

        if operation == "add":
            if not is_always_ready_stock(spare_part.stok):
                spare_part.stok += quantity
        elif operation == "subtract":
            if not is_always_ready_stock(spare_part.stok):
                if spare_part.stok < quantity:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Stok tidak mencukupi. Stok tersedia: {spare_part.stok}",
                    )
                spare_part.stok -= quantity
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Operation harus 'add' atau 'subtract'",
            )

        self.db.commit()
        self.db.refresh(spare_part)
        self._emit_change("stock_updated", spare_part)

        return spare_part

    def update_price(
        self,
        spare_part_id: int,
        harga_beli: Optional[Decimal] = None,
        harga_jual: Optional[Decimal] = None,
    ) -> SparePart:
        """Update spare part prices."""
        spare_part = self.get_by_id(spare_part_id)

        if harga_beli is not None:
            spare_part.harga_beli = harga_beli
        if harga_jual is not None:
            spare_part.harga_jual = harga_jual

        self.db.commit()
        self.db.refresh(spare_part)
        self._emit_change("price_updated", spare_part)

        return spare_part

    def get_low_stock_items(self) -> List[SparePart]:
        """Get spare parts with low stock."""
        return (
            self.db.query(SparePart)
            .filter(
                SparePart.deleted_at.is_(None),
                SparePart.stok <= SparePart.stok_minimum,
                SparePart.stok != ALWAYS_READY_STOCK,
            )
            .order_by(SparePart.stok.asc())
            .all()
        )

    def get_stock_value(self) -> Dict[str, Any]:
        """Get total stock value."""
        # For normal items: modal = stok × harga_beli
        # For Always Ready (999999): modal = 0 (catalog reference only, no physical stock)
        result = (
            self.db.query(
                func.sum(
                    case(
                        (SparePart.stok == ALWAYS_READY_STOCK, 0),
                        else_=SparePart.stok * SparePart.harga_beli
                    )
                ).label("total_value"),
                func.sum(
                    case(
                        (SparePart.stok != ALWAYS_READY_STOCK, SparePart.stok),
                        else_=0
                    )
                ).label("total_items"),
                func.count(SparePart.id).label("total_products"),
            )
            .filter(SparePart.deleted_at.is_(None))
            .first()
        )

        return {
            "total_value": float(result.total_value or 0),
            "total_items": int(result.total_items or 0),
            "total_products": int(result.total_products or 0),
        }

    def search_for_transaction(
        self,
        query: str,
        limit: int = 20,
    ) -> List[SparePart]:
        """Search spare parts for transaction form."""
        search_filter = f"%{query}%"
        return (
            self.db.query(SparePart)
            .filter(
                SparePart.deleted_at.is_(None),
                SparePart.stok > 0,
                or_(
                    SparePart.nama.ilike(search_filter),
                    SparePart.kode.ilike(search_filter),
                    SparePart.kode_part.ilike(search_filter),
                    SparePart.kode_ean.ilike(search_filter),
                ),
            )
            .order_by(SparePart.nama.asc())
            .limit(limit)
            .all()
        )

    def get_categories(self) -> List[str]:
        """Get distinct categories."""
        categories = (
            self.db.query(SparePart.kategori)
            .filter(
                SparePart.deleted_at.is_(None),
                SparePart.kategori.isnot(None),
            )
            .distinct()
            .all()
        )
        return [c[0] for c in categories if c[0]]

    def get_brands(self) -> List[str]:
        """Get distinct brands."""
        brands = (
            self.db.query(SparePart.merek)
            .filter(
                SparePart.deleted_at.is_(None),
                SparePart.merek.isnot(None),
            )
            .distinct()
            .all()
        )
        return [b[0] for b in brands if b[0]]
    
    def get_stats(self) -> Dict[str, Any]:
        """Get inventory statistics (top sales and lowest stock)."""
        from app.models.bengkel import DetailTransaksiSpareParts
        
        # 1. Top 5 Sales
        top_sales = (
            self.db.query(
                SparePart,
                func.sum(DetailTransaksiSpareParts.qty).label("total_sales")
            )
            .join(DetailTransaksiSpareParts)
            .filter(SparePart.deleted_at.is_(None))
            .group_by(SparePart.id)
            .order_by(func.sum(DetailTransaksiSpareParts.qty).desc())
            .limit(5)
            .all()
        )
        
        # Format top_sales to include total_sales in response
        top_sales_formatted = []
        for part, sales in top_sales:
            # We can't easily add attributes to SQLAlchemy models, so we'll return a dict or similar
            # But let's just use the model and let the speaker handle the rest if needed
            # For simplicity, we'll return the part object and add an extra field in a dict
            top_sales_formatted.append({
                "id": part.id,
                "nama": part.nama,
                "kode": part.kode,
                "stok": part.stok,
                "stok_minimum": part.stok_minimum,
                "harga_jual": float(part.harga_jual or 0),
                "kategori": part.kategori,
                "satuan": part.satuan,
                "total_sales": int(sales or 0),
                "gambar": part.gambar
            })
            
        # 2. Top 5 Lowest Stock (Excluding 999)
        lowest_stock = (
            self.db.query(SparePart)
            .filter(
                SparePart.deleted_at.is_(None),
                SparePart.stok != ALWAYS_READY_STOCK
            )
            .order_by(SparePart.stok.asc())
            .limit(5)
            .all()
        )
        
        lowest_stock_formatted = []
        for part in lowest_stock:
            lowest_stock_formatted.append({
                "id": part.id,
                "nama": part.nama,
                "kode": part.kode,
                "stok": part.stok,
                "stok_minimum": part.stok_minimum,
                "harga_jual": float(part.harga_jual or 0),
                "kategori": part.kategori,
                "satuan": part.satuan,
                "gambar": part.gambar
            })
            
        return {
            "top_sales": top_sales_formatted,
            "lowest_stock": lowest_stock_formatted
        }

    @staticmethod
    def _is_tanpa_stok_marker(stok_raw) -> bool:
        """True when the Stok cell explicitly means no physical stock (Always Ready)."""
        if stok_raw is None:
            return False
        stok_str = str(stok_raw).strip().lower()
        if not stok_str:
            return False
        if stok_str in ('tanpa stok', 'always ready', 'ar', 'n/a', 'na', '-', '—'):
            return True
        return 'tanpa' in stok_str and 'stok' in stok_str

    def _detect_format(self, sheet) -> str:
        """Detect Excel import format based on header row.
        
        Returns:
            'stok_format' - import-stok-format.xlsx style (Urutan, Nama, Kode Part, Harga Beli, Harga Jual, Stok, Satuan, Total Modal)
            'standard'    - Original 12-column format (Kode, Nama, Kode Part, Kategori, Merek, Satuan, Stok, Stok Min, Harga Beli, Harga Jual, Lokasi Rak, Catatan)
        """
        header_row = [str(cell.value or '').strip().lower() for cell in sheet[1]]
        
        # Check for import-stok-format pattern
        col_a = header_row[0] if len(header_row) > 0 else ''
        col_b = header_row[1] if len(header_row) > 1 else ''
        
        stok_format_indicators = ['urutan', 'no', 'no.', 'nomor', 'urutan sparepart']
        if any(indicator in col_a for indicator in stok_format_indicators):
            return 'stok_format'
        
        # Also detect by checking if col D header looks like "Harga Beli" (stok format)
        # vs "Kategori" (standard format)
        col_d = header_row[3] if len(header_row) > 3 else ''
        if 'harga' in col_d and 'beli' in col_d:
            return 'stok_format'
        
        return 'standard'

    def _parse_row_stok_format(self, row, row_idx: int) -> Dict[str, Any]:
        """Parse a row from import-stok-format.xlsx.
        
        Columns:
        A(0): Urutan Sparepart (ignored, just a sequence number)
        B(1): Nama Spare Part (Required)
        C(2): Kode Part
        D(3): Harga Beli
        E(4): Harga Jual
        F(5): Stok
        G(6): Satuan
        H(7): Total Modal (used to derive exact stok for normal items)
        I(8): Always Ready (optional, true/ya/yes/1 = stok 999999, modal=0)
        K(10): Total Fix (validation total, read separately)
        """
        nama = str(row[1]).strip() if row[1] else None
        if not nama:
            return {"error": f"Baris {row_idx}: Nama spare part wajib diisi"}
        
        try:
            harga_beli = Decimal(str(row[3] or 0))
            harga_jual = Decimal(str(row[4] or 0))

            stok = None

            # Detect Always Ready marker:
            # 1. Explicit col I marker (true/ya/yes/1)
            # 2. Stok column contains text like "Tanpa Stok"
            always_ready = False
            if len(row) > 8 and row[8] is not None:
                ar_val = str(row[8]).strip().lower()
                always_ready = ar_val in ('true', 'ya', 'yes', '1', 'v', '✓', 'always ready')

            stok_raw = row[5]
            stok_str = str(stok_raw or '').strip()
            stok_is_tanpa_stok = self._is_tanpa_stok_marker(stok_raw)
            if not always_ready and stok_is_tanpa_stok:
                always_ready = True
            
            # --- Read Excel's "Total Modal" (col H) as source of truth ---
            excel_modal = Decimal("0")
            if len(row) > 7 and row[7]:
                try:
                    excel_modal = Decimal(str(row[7]).strip())
                except InvalidOperation:
                    pass
            
            if always_ready:
                if stok_is_tanpa_stok:
                    # "Tanpa Stok" in kolom Stok → selalu katalog (stok 999999)
                    stok = ALWAYS_READY_STOCK
                elif excel_modal > 0 and harga_beli > 0:
                    # AR via kolom I + Total Modal: derive stok fisik dari modal
                    stok = excel_modal / harga_beli
                else:
                    stok = ALWAYS_READY_STOCK
            else:
                # Normal item: use Total Modal to derive exact stok if available
                if excel_modal > 0 and harga_beli > 0:
                    stok = excel_modal / harga_beli
                elif stok_raw is not None:
                    try:
                        stok = Decimal(stok_str) if stok_str else Decimal("0")
                    except InvalidOperation:
                        stok = Decimal("0")
                else:
                    stok = Decimal("0")
            
            satuan = str(row[6]).strip() if row[6] and str(row[6]).strip() else "pcs"
            kode_part = str(row[2]).strip() if row[2] and str(row[2]).strip() else None
            
            return {
                "kode": None,  # No internal kode in this format, will be auto-generated
                "nama": nama,
                "kode_part": kode_part,
                "kategori": "Umum",
                "merek": None,
                "satuan": satuan,
                "stok": stok,
                "stok_minimum": 5,
                "harga_beli": harga_beli,
                "harga_jual": harga_jual,
                "lokasi_rak": None,
                "catatan": None,
            }
        except (ValueError, TypeError, InvalidOperation) as e:
            return {"error": f"Baris {row_idx}: Format data numerik tidak valid ({str(e)})"}

    def _parse_row_standard(self, row, row_idx: int) -> Dict[str, Any]:
        """Parse a row from the standard 12-column export format.
        
        Columns:
        A(0): Kode Barang (Internal)
        B(1): Nama (Required)
        C(2): Kode Part (OEM / Mfg Code)
        D(3): Kategori
        E(4): Merek
        F(5): Satuan
        G(6): Stok
        H(7): Stok Minimum
        I(8): Harga Beli
        J(9): Harga Jual
        K(10): Lokasi Rak
        L(11): Catatan
        """
        kode = str(row[0]).strip() if row[0] else None
        nama = str(row[1]).strip() if row[1] else None
        
        if not nama:
            return {"error": f"Baris {row_idx}: Nama spare part wajib diisi"}
        
        try:
            stok_raw = row[6]
            if self._is_tanpa_stok_marker(stok_raw):
                # Always Ready / katalog tanpa stok fisik
                stok = ALWAYS_READY_STOCK
            elif stok_raw is None or str(stok_raw).strip() == "":
                stok = Decimal("0")
            else:
                # Numeric(15,2) — jaga stok desimal (0.5, 3.3) agar grand modal tetap akurat
                stok = Decimal(str(stok_raw).strip().replace(",", "."))
            harga_beli = Decimal(str(row[8] or 0))
            
            return {
                "kode": kode,
                "nama": nama,
                "kode_part": str(row[2]).strip() if row[2] else None,
                "kode_ean": str(row[12]).strip() if len(row) > 12 and row[12] else None,
                "kategori": str(row[3]) if row[3] else "Umum",
                "merek": str(row[4]) if row[4] else None,
                "satuan": str(row[5]) if row[5] else "pcs",
                "stok": stok,
                "stok_minimum": int(row[7]) if row[7] is not None else 5,
                "harga_beli": harga_beli,
                "harga_jual": Decimal(str(row[9] or 0)),
                "lokasi_rak": str(row[10]) if len(row) > 10 and row[10] else None,
                "catatan": str(row[11]) if len(row) > 11 and row[11] else None,
            }
        except (ValueError, TypeError, InvalidOperation) as e:
            return {"error": f"Baris {row_idx}: Format data numerik tidak valid ({str(e)})"}

    def import_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Import spare parts from Excel file.
        
        Auto-detects two formats:
        
        FORMAT 1 - 'stok_format' (import-stok-format.xlsx):
        A: Urutan Sparepart (No.)
        B: Nama Spare Part (Required)
        C: Kode Part
        D: Harga Beli
        E: Harga Jual
        F: Stok
        G: Satuan
        H: Total Modal (ignored)
        
        FORMAT 2 - 'standard' (export/original format):
        A: Kode Barang (Internal)
        B: Nama (Required)
        C: Kode Part (OEM / Mfg Code)
        D: Kategori
        E: Merek
        F: Satuan
        G: Stok
        H: Stok Minimum
        I: Harga Beli
        J: Harga Jual
        K: Lokasi Rak
        L: Catatan
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Gagal membaca file Excel: {str(e)}"
            )

        # Auto-detect format
        detected_format = self._detect_format(sheet)

        # Read Total Fix from column K (index 10) if present (stok_format)
        total_fix_excel = None
        if detected_format == 'stok_format':
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 10 and row[10] is not None:
                    try:
                        total_fix_excel = float(row[10])
                    except (ValueError, TypeError):
                        pass
                    break

        results = {
            "total": 0,
            "success": 0,
            "updated": 0,
            "duplicates": 0,
            "failed": 0,
            "skipped": 0,
            "errors": [],
            "format_detected": detected_format,
        }
        
        # ===================================================================
        # PHASE 1: Parse all rows into in-memory dicts (no DB operations)
        # ===================================================================
        parsed_rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): 
                results["skipped"] += 1
                continue
            
            results["total"] += 1

            if detected_format == 'stok_format':
                parsed = self._parse_row_stok_format(row, row_idx)
            else:
                parsed = self._parse_row_standard(row, row_idx)

            if "error" in parsed:
                results["failed"] += 1
                results["errors"].append(parsed["error"])
                continue
            
            parsed["_row_idx"] = row_idx
            parsed_rows.append(parsed)

        # ===================================================================
        # PHASE 2: Apply parsed rows to DB (hard delete + fresh insert)
        # ===================================================================
        try:
            # Step 1: HARD DELETE all existing spare parts for a truly fresh import.
            from sqlalchemy import text
            self.db.execute(text("SET FOREIGN_KEY_CHECKS = 0"))
            self.db.execute(
                SparePart.__table__.delete()
            )
            self.db.execute(text("SET FOREIGN_KEY_CHECKS = 1"))
            self.db.flush()

            # Step 2: Pre-generate kodes for items that need one.
            date_str = datetime.now().strftime("%y%m")
            kodes_needed = sum(1 for p in parsed_rows if not p.get("kode"))
            pre_generated_kodes = [f"SPR{date_str}{(i + 1):04d}" for i in range(kodes_needed)]
            kode_idx = 0

            # Step 3: Insert all rows WITHOUT merging (Patokan = Urutan di Excel)
            for parsed in parsed_rows:
                row_idx = parsed["_row_idx"]
                nama = parsed["nama"]
                stok = parsed["stok"]
                harga_beli = parsed["harga_beli"]

                try:
                    # --- CREATE new item ---
                    kode = parsed.get("kode")
                    if kode:
                        new_kode = kode
                    else:
                        new_kode = pre_generated_kodes[kode_idx]
                        kode_idx += 1
                    
                    new_spare_part = SparePart(
                        kode=new_kode,
                        nama=nama,
                        kode_part=parsed["kode_part"],
                        kode_ean=parsed.get("kode_ean"),
                        kategori=parsed["kategori"] or "Umum",
                        merek=parsed["merek"],
                        satuan=parsed["satuan"] or "pcs",
                        stok=stok,
                        stok_minimum=parsed.get("stok_minimum", 5),
                        harga_beli=harga_beli,
                        harga_jual=parsed["harga_jual"],
                        lokasi_rak=parsed["lokasi_rak"],
                        catatan=parsed["catatan"],
                    )
                    self.db.add(new_spare_part)
                    results["success"] += 1

                except Exception as e:
                    results["failed"] += 1
                    results["errors"].append(f"Baris {row_idx}: {str(e)}")
                    continue

            # Single flush + commit for the entire batch
            self.db.flush()
            self.db.commit()
        except Exception as e:
            self.db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Import gagal: {str(e)}"
            )
        
        # ===================================================================
        # PHASE 3: Post-import validation — compare DB modal vs Total Fix
        # ===================================================================
        stock_value = self.get_stock_value()
        db_total_modal = stock_value["total_value"]
        results["total_modal_db"] = db_total_modal
        
        if total_fix_excel is not None:
            results["total_fix_excel"] = total_fix_excel
            diff = abs(db_total_modal - total_fix_excel)
            results["modal_diff"] = diff
            if diff >= 1:
                results["modal_warning"] = (
                    f"⚠️ Total modal di database (Rp {db_total_modal:,.0f}) "
                    f"tidak sesuai dengan Total Fix di Excel (Rp {total_fix_excel:,.0f}). "
                    f"Selisih: Rp {diff:,.0f}"
                )
            else:
                results["modal_verified"] = True
        
        return results

    def bulk_delete(self, ids: List[int]) -> int:
        """Deletes multiple spare parts by ID (soft delete)."""
        now = datetime.now()
        updated = (
            self.db.query(SparePart)
            .filter(SparePart.id.in_(ids))
            .filter(SparePart.deleted_at.is_(None))
            .update({SparePart.deleted_at: now}, synchronize_session=False)
        )
        self.db.commit()
        self._emit_change("bulk_deleted")
        return updated

    def export_to_excel(self, ids: Optional[List[int]] = None) -> io.BytesIO:
        """Export non-deleted spare parts to Excel format."""
        query = self.db.query(SparePart).filter(SparePart.deleted_at.is_(None))
        if ids:
            query = query.filter(SparePart.id.in_(ids))
            
        spare_parts = query.all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spare Parts"
        
        # Headers
        headers = [
            "Kode", "Nama Barang", "Kode Part", "Kategori", "Merek",
            "Satuan", "Stok", "Stok Minimal", "Harga Beli", "Harga Jual",
            "Lokasi Rak", "Catatan", "Kode EAN"
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            
        # Data
        for row_idx, sp in enumerate(spare_parts, 2):
            ws.cell(row=row_idx, column=1, value=sp.kode)
            ws.cell(row=row_idx, column=2, value=sp.nama)
            ws.cell(row=row_idx, column=3, value=sp.kode_part)
            ws.cell(row=row_idx, column=4, value=sp.kategori)
            ws.cell(row=row_idx, column=5, value=sp.merek)
            ws.cell(row=row_idx, column=6, value=sp.satuan)
            ws.cell(row=row_idx, column=7, value=sp.stok)
            ws.cell(row=row_idx, column=8, value=sp.stok_minimum)
            ws.cell(row=row_idx, column=9, value=float(sp.harga_beli) if sp.harga_beli else 0)
            ws.cell(row=row_idx, column=10, value=float(sp.harga_jual) if sp.harga_jual else 0)
            ws.cell(row=row_idx, column=11, value=sp.lokasi_rak)
            ws.cell(row=row_idx, column=12, value=sp.catatan)
            ws.cell(row=row_idx, column=13, value=sp.kode_ean)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_import_template_instructions(self, ws, lines: List[str]) -> None:
        ws.column_dimensions["A"].width = 96
        title_font = openpyxl.styles.Font(bold=True, size=13, color="1D4ED8")
        heading_font = openpyxl.styles.Font(bold=True, size=11, color="111827")
        body_font = openpyxl.styles.Font(size=11, color="374151")
        wrap = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        for row_idx, line in enumerate(lines, start=1):
            cell = ws.cell(row=row_idx, column=1, value=line)
            cell.alignment = wrap
            if row_idx == 1:
                cell.font = title_font
            elif line.endswith(":") or line.isupper():
                cell.font = heading_font
            else:
                cell.font = body_font

    def export_import_template(self, format_type: str = "stok_format") -> io.BytesIO:
        """Generate Excel import template with sample data sheet + instruction sheet."""
        wb = openpyxl.Workbook()
        ws_data = wb.active

        header_fill = openpyxl.styles.PatternFill("solid", fgColor="E0E7FF")
        header_font = openpyxl.styles.Font(bold=True, color="1E3A8A")
        instruction_lines: List[str] = []

        if format_type == "standard":
            ws_data.title = "Import Standar"
            instruction_lines = [
                "PETUNJUK IMPORT SPAREPART - FORMAT STANDAR",
                "",
                "LANGKAH PENGGUNAAN:",
                "1. Isi data pada sheet 'Import Standar' mulai baris 2.",
                "2. Baris 1 adalah header — jangan diubah atau dihapus.",
                "3. Baris 2 berisi contoh — hapus atau ganti dengan data Anda.",
                "4. Simpan file, lalu upload melalui menu Import di aplikasi TPM.",
                "5. Import akan MENGGANTI seluruh data sparepart yang ada.",
                "",
                "PENJELASAN KOLOM:",
                "A - Kode: kode internal TPM. Kosongkan untuk auto-generate (SPRyyMM0001).",
                "B - Nama Barang: wajib diisi.",
                "C - Kode Part: kode pabrik / GS1 (90), contoh MD273133003700079.",
                "D - Kategori: default Umum jika kosong.",
                "E - Merek: merek barang (opsional).",
                "F - Satuan: pcs, liter, set, dll. Default pcs.",
                "G - Stok: jumlah stok saat ini.",
                "H - Stok Minimal: batas peringatan stok rendah. Default 5.",
                "I - Harga Beli: harga modal per satuan (angka).",
                "J - Harga Jual: harga jual per satuan (angka).",
                "K - Lokasi Rak: posisi rak gudang (opsional).",
                "L - Catatan: catatan tambahan (opsional).",
                "M - Kode EAN: barcode EAN-13 untuk scan, contoh 8996001326398.",
                "",
                "TIPS UPDATE DATA:",
                "• Export data existing dulu, edit di Excel, lalu import ulang.",
                "• Isi Kode Part dan Kode EAN agar scan barcode/QR berfungsi.",
                "• Baris kosong otomatis dilewati saat import.",
            ]
            headers = [
                "Kode", "Nama Barang", "Kode Part", "Kategori", "Merek",
                "Satuan", "Stok", "Stok Minimal", "Harga Beli", "Harga Jual",
                "Lokasi Rak", "Catatan", "Kode EAN",
            ]
            sample = [
                "SPR25010001", "Contoh Oli Mesin", "MD273133003700079", "Umum", "Castrol",
                "pcs", 10, 5, 50000, 75000, "Rak A1", "Contoh catatan", "8996001326398",
            ]
        else:
            ws_data.title = "Import Stok"
            instruction_lines = [
                "PETUNJUK IMPORT SPAREPART - FORMAT STOK",
                "",
                "LANGKAH PENGGUNAAN:",
                "1. Isi data pada sheet 'Import Stok' mulai baris 2.",
                "2. Baris 1 adalah header — jangan diubah atau dihapus.",
                "3. Baris 2 berisi contoh — hapus atau ganti dengan data Anda.",
                "4. Simpan file, lalu upload melalui menu Import di aplikasi TPM.",
                "5. Import akan MENGGANTI seluruh data sparepart yang ada.",
                "",
                "PENJELASAN KOLOM:",
                "A - Urutan Sparepart: nomor urut (informasi saja).",
                "B - Nama Spare Part: wajib diisi.",
                "C - Kode Part: kode pabrik / GS1 (opsional).",
                "D - Harga Beli: harga modal per satuan (angka).",
                "E - Harga Jual: harga jual per satuan (angka).",
                "F - Stok: angka, atau teks 'Tanpa Stok' (otomatis Always Ready, stok 999999).",
                "G - Satuan: pcs, liter, set, dll. Default pcs.",
                "H - Total Modal: harga beli × stok. Dipakai untuk hitung stok akurat.",
                "I - Always Ready: isi ya / true / 1 (opsional jika kolom Stok sudah 'Tanpa Stok').",
                "K - Total Fix: total modal keseluruhan untuk validasi (opsional).",
                "",
                "CATATAN ALWAYS READY:",
                "• Kolom Stok berisi 'Tanpa Stok' → sistem otomatis set stok 999999.",
                "• Kolom I + Total Modal > 0 → stok dihitung dari modal (barang AR dengan stok fisik).",
                "• Always Ready tanpa modal → stok 999999 (katalog saja).",
                "",
                "TIPS:",
                "• Pastikan Total Modal = Harga Beli × Stok untuk barang biasa.",
                "• Baris kosong otomatis dilewati saat import.",
            ]
            headers = [
                "Urutan Sparepart", "Nama Spare Part", "Kode Part",
                "Harga Beli", "Harga Jual", "Stok", "Satuan",
                "Total Modal", "Always Ready", "", "Total Fix",
            ]
            sample = [
                1, "Contoh Oli Mesin", "MD273133003700079",
                50000, 75000, 10, "pcs", 500000, "", "", 500000,
            ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for col_idx, value in enumerate(sample, 1):
            ws_data.cell(row=2, column=col_idx, value=value)

        ws_data.freeze_panes = "A2"

        ws_info = wb.create_sheet("Instruksi")
        self._write_import_template_instructions(ws_info, instruction_lines)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
