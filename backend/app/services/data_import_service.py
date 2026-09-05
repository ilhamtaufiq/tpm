"""
Multi-sheet Excel import for existing/opening data.

Sheets (MVP):
  _INSTRUKSI, customers, suppliers, spare_parts, jasa_servis, karyawan,
  kas_opening, hutang_opening, piutang_opening, armada, supir, mobil

Flow:
  1) generate_template()
  2) preview(file) dry-run â€” no commit
  3) commit(file, batch_id?) â€” all-or-nothing transaction
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.models.customer import Customer
from app.models.supplier import Supplier
from app.models.bengkel import SparePart, JasaServis
from app.models.karyawan import Karyawan
from app.models.keuangan import KasBank, HutangUsaha, PiutangUsaha, Aset
from app.models.jasa_angkut import ArmadaJasaAngkut, Supir
from app.models.mobil import Mobil
from app.utils.constants import (
    CarStatus,
    EmployeeStatus,
    HutangSource,
    HutangStatus,
    KasBankJenis,
    KasBankSource,
    KasBankType,
    OwnershipType,
    PaymentMethod,
    PaymentStatus,
    PiutangSource,
    PiutangStatus,
    AssetCategory,
    AssetStatus,
    TRANSACTION_PREFIXES,
)

SHEET_ORDER = [
    "_INSTRUKSI",
    "customers",
    "suppliers",
    "spare_parts",
    "jasa_servis",
    "karyawan",
    "kas_opening",
    "asset",
    "hutang_opening",
    "piutang_opening",
    "armada",
    "supir",
    "mobil",
    "neraca_check",
]

# Header definitions: (excel_header, field_key, required)
SHEET_HEADERS: Dict[str, List[Tuple[str, str, bool]]] = {
    "customers": [
        ("kode", "kode", False),
        ("nama", "nama", True),
        ("tipe", "tipe", False),
        ("telepon", "telepon", False),
        ("alamat", "alamat", False),
        ("kota", "kota", False),
        ("email", "email", False),
        ("npwp", "npwp", False),
        ("catatan", "catatan", False),
    ],
    "suppliers": [
        ("kode", "kode", False),
        ("nama", "nama", True),
        ("telepon", "telepon", False),
        ("alamat", "alamat", False),
        ("kota", "kota", False),
        ("email", "email", False),
        ("contact_person", "contact_person", False),
        ("bank", "bank", False),
        ("rekening", "rekening", False),
        ("catatan", "catatan", False),
    ],
    "spare_parts": [
        ("kode", "kode", False),
        ("nama", "nama", True),
        ("kode_part", "kode_part", False),
        ("kategori", "kategori", False),
        ("merek", "merek", False),
        ("satuan", "satuan", False),
        ("stok", "stok", False),
        ("stok_minimum", "stok_minimum", False),
        ("harga_beli", "harga_beli", False),
        ("harga_jual", "harga_jual", False),
        ("lokasi_rak", "lokasi_rak", False),
        ("catatan", "catatan", False),
    ],
    "jasa_servis": [
        ("nama", "nama", True),
        ("kategori", "kategori", False),
        ("harga", "harga", True),
        ("deskripsi", "deskripsi", False),
    ],
    "karyawan": [
        ("kode", "kode", False),
        ("nama", "nama", True),
        ("jabatan", "jabatan", True),
        ("nik", "nik", False),
        ("telepon", "telepon", False),
        ("alamat", "alamat", False),
        ("email", "email", False),
        ("tanggal_bergabung", "tanggal_bergabung", False),
        ("gaji_pokok", "gaji_pokok", False),
        ("tunjangan", "tunjangan", False),
        ("status", "status", False),
        ("catatan", "catatan", False),
    ],
    "kas_opening": [
        ("tanggal", "tanggal", True),
        ("jenis_kas", "jenis_kas", True),
        ("nominal", "nominal", True),
        ("keterangan", "keterangan", False),
        ("catatan", "catatan", False),
    ],
    "asset": [
        ("kode", "kode", False),
        ("tanggal", "tanggal", False),
        ("nama asset", "nama", True),
        ("nominal", "harga_beli", True),
        ("kategori", "kategori", False),
        ("catatan", "catatan", False),
    ],
    "hutang_opening": [
        ("tanggal", "tanggal", True),
        ("nama_kreditur", "nama_kreditur", True),
        ("nominal", "nominal", True),
        ("unit", "unit", False),
        ("telepon", "telepon", False),
        ("alamat", "alamat", False),
        ("tanggal_jatuh_tempo", "tanggal_jatuh_tempo", False),
        ("catatan", "catatan", False),
    ],
    "piutang_opening": [
        ("tanggal", "tanggal", True),
        ("nama_debitur", "nama_debitur", True),
        ("nominal", "nominal", True),
        ("unit", "unit", False),
        ("telepon", "telepon", False),
        ("alamat", "alamat", False),
        ("tanggal_jatuh_tempo", "tanggal_jatuh_tempo", False),
        ("catatan", "catatan", False),
    ],
    "armada": [
        ("nama", "nama", True),
        ("nopol", "nopol", True),
        ("jenis", "jenis", False),
        ("is_active", "is_active", False),
        ("catatan", "catatan", False),
    ],
    "supir": [
        ("kode", "kode", False),
        ("nama", "nama", True),
        ("telepon", "telepon", False),
        ("nik", "nik", False),
        ("nomor_sim", "nomor_sim", False),
        ("jenis_sim", "jenis_sim", False),
        ("tanggal_bergabung", "tanggal_bergabung", False),
        ("nopol_kendaraan", "nopol_kendaraan", False),
        ("is_active", "is_active", False),
        ("catatan", "catatan", False),
    ],
    "mobil": [
        ("kode", "kode", False),
        ("merek", "merek", True),
        ("model", "model", True),
        ("tahun", "tahun", True),
        ("warna", "warna", True),
        ("nomor_plat", "nomor_plat", True),
        ("no mesin", "nomor_mesin", False),
        ("no rangka", "nomor_rangka", False),
        ("harga_beli", "harga_beli", True),
        ("harga_jual", "harga_jual", False),
        ("tanggal_masuk", "tanggal_masuk", False),
        ("status", "status", False),
        ("tipe_kepemilikan", "tipe_kepemilikan", False),
        ("nama_investor", "nama_investor", False),
        ("nominal investasi", "nominal_investor", False),
        ("biaya pengeluaran/ operasional", "biaya_ops", False),
        ("keterangan biaya pengeluaran/ operasional", "biaya_ops_ket", False),
        ("biaya part dan service", "biaya_part_service", False),
        ("transmisi", "transmisi", False),
        ("bahan_bakar", "bahan_bakar", False),
        ("kilometer", "kilometer", False),
        ("catatan", "catatan", False),
    ],
    "neraca_check": [
        # === AKTIVA LANCAR: Kas & Bank ===
        ("kas_tunai", "kas_tunai", False),
        ("kas_bank", "kas_bank", False),
        ("unit_cash", "unit_cash", False),
        ("total_kas_bank", "total_kas_bank", False),
        # === AKTIVA LANCAR: Piutang Usaha ===
        ("piutang_lainnya", "piutang_lainnya", False),
        ("piutang_karyawan", "piutang_karyawan", False),
        ("piutang_usaha", "piutang_usaha", False),
        ("piutang_mobil", "piutang_mobil", False),
        ("piutang_jasa_angkut", "piutang_jasa_angkut", False),
        ("total_piutang", "total_piutang", False),
        # === AKTIVA LANCAR: Persediaan & Stok ===
        ("persediaan_sparepart", "persediaan_sparepart", False),
        ("stok_mobil", "stok_mobil", False),
        ("stok_mobil_detail_harga_beli", "stok_mobil_detail_harga_beli", False),
        ("stok_mobil_detail_biaya_persiapan", "stok_mobil_detail_biaya_persiapan", False),
        ("stok_mobil_detail_perbaikan_external", "stok_mobil_detail_perbaikan_external", False),
        ("stok_mobil_detail_perbaikan_internal", "stok_mobil_detail_perbaikan_internal", False),
        ("total_aktiva_lancar", "total_aktiva_lancar", False),
        # === AKTIVA TETAP: Daftar Aset Aktif ===
        ("detail_aset_kode", "detail_aset_kode", False),
        ("detail_aset_nama", "detail_aset_nama", False),
        ("detail_aset_harga_beli", "detail_aset_harga_beli", False),
        ("total_aset_tetap", "total_aset_tetap", False),
        ("total_aktiva", "total_aktiva", False),
        # === PASIVA: Modal ===
        ("setoran_modal", "setoran_modal", False),
        ("setoran_modal_kas", "setoran_modal_kas", False),
        ("modal_non_kas", "modal_non_kas", False),
        ("modal_persediaan", "modal_persediaan", False),
        ("modal_stok_mobil", "modal_stok_mobil", False),
        ("modal_aset_tetap", "modal_aset_tetap", False),
        ("laba_ditahan", "laba_ditahan", False),
        ("prive", "prive", False),
        ("total_modal", "total_modal", False),
        # === PASIVA: Hutang ===
        ("hutang_part", "hutang_part", False),
        ("hutang_mobil", "hutang_mobil", False),
        ("hutang_investor", "hutang_investor", False),
        ("hutang_lainnya", "hutang_lainnya", False),
        ("hutang_jasa_angkut", "hutang_jasa_angkut", False),
        ("uang_muka_penjualan", "uang_muka_penjualan", False),
        ("piutang_booking", "piutang_booking", False),
        ("total_hutang", "total_hutang", False),
        # === BALANCE CHECK ===
        ("total_pasiva", "total_pasiva", False),
        ("selisih", "selisih", False),
    ],
}

EXAMPLE_ROWS: Dict[str, List[Dict[str, Any]]] = {
    "customers": [
        {
            "kode": "CUS-001",
            "nama": "Budi Santoso",
            "tipe": "perorangan",
            "telepon": "08123456789",
            "alamat": "Jl. Contoh 1",
            "kota": "Cianjur",
        },
        {
            "kode": "CUS-002",
            "nama": "PT Maju Jaya",
            "tipe": "perusahaan",
            "telepon": "081398765432",
            "alamat": "Jl. Raya Bandung No. 10",
            "kota": "Bandung",
            "email": "admin@majujaya.co.id",
        },
    ],
    "suppliers": [
        {
            "kode": "SUP-001",
            "nama": "Toko Onderdil Jaya",
            "telepon": "0811111111",
            "kota": "Bandung",
        },
        {
            "kode": "SUP-002",
            "nama": "CV Sumber Sparepart",
            "telepon": "08223334444",
            "kota": "Jakarta",
            "contact_person": "Ibu Sari",
        },
    ],
    "spare_parts": [
        {
            "kode": "SPR-001",
            "nama": "Oli Mesin 1L",
            "kode_part": "OLI-001",
            "kategori": "Oli",
            "satuan": "pcs",
            "stok": 50,
            "stok_minimum": 5,
            "harga_beli": 35000,
            "harga_jual": 45000,
        },
        {
            "kode": "SPR-002",
            "nama": "Filter Udara Avanza",
            "kode_part": "FLT-002",
            "kategori": "Filter",
            "satuan": "pcs",
            "stok": 20,
            "stok_minimum": 3,
            "harga_beli": 75000,
            "harga_jual": 95000,
        },
        {
            "kode": "SPR-003",
            "nama": "Kampas Rem Depan",
            "kode_part": "BRK-003",
            "kategori": "Rem",
            "satuan": "set",
            "stok": 10,
            "stok_minimum": 2,
            "harga_beli": 120000,
            "harga_jual": 150000,
        },
    ],
    "jasa_servis": [
        {"nama": "Ganti Oli", "kategori": "Mesin", "harga": 50000, "deskripsi": "Jasa ganti oli"},
        {"nama": "Servis Rutin", "kategori": "Perawatan", "harga": 150000, "deskripsi": "Servis berkala 10.000 km"},
        {"nama": "Ganti Kampas Rem", "kategori": "Rem", "harga": 100000, "deskripsi": "Jasa ganti kampas rem"},
    ],
    "karyawan": [
        {
            "kode": "KRY-001",
            "nama": "Andi Mekanik",
            "jabatan": "Mekanik",
            "tanggal_bergabung": date.today().isoformat(),
            "gaji_pokok": 3000000,
            "tunjangan": 500000,
            "status": "AKTIF",
        },
        {
            "kode": "KRY-002",
            "nama": "Siti Admin",
            "jabatan": "Admin",
            "tanggal_bergabung": date.today().isoformat(),
            "gaji_pokok": 2500000,
            "tunjangan": 300000,
            "status": "AKTIF",
        },
    ],
    "kas_opening": [
        {
            "tanggal": date.today().isoformat(),
            "jenis_kas": "KAS_UTAMA",
            "nominal": 10000000,
            "keterangan": "Saldo awal kas utama",
        },
        {
            "tanggal": date.today().isoformat(),
            "jenis_kas": "BANK_UTAMA",
            "nominal": 50000000,
            "keterangan": "Saldo awal bank utama (BCA)",
        },
        {
            "tanggal": date.today().isoformat(),
            "jenis_kas": "KAS_UNIT_BENGKEL",
            "nominal": 2000000,
            "keterangan": "Saldo awal laci bengkel",
        },
        {
            "tanggal": date.today().isoformat(),
            "jenis_kas": "KAS_UNIT_JASA_ANGKUT",
            "nominal": 1500000,
            "keterangan": "Saldo awal kas unit jasa angkut",
        },
        {
            "tanggal": date.today().isoformat(),
            "jenis_kas": "KAS_UNIT_MOBIL",
            "nominal": 3000000,
            "keterangan": "Saldo awal kas unit jual-beli mobil",
        },
    ],
    "asset": [
        {
            "tanggal": date.today().isoformat(),
            "nama": "Laptop Asus",
            "harga_beli": 8000000,
            "kategori": "ELECTRONIC",
            "catatan": "Aset kantor",
        },
        {
            "tanggal": date.today().isoformat(),
            "nama": "Meja Kantor",
            "harga_beli": 1500000,
            "kategori": "LAINNYA",
            "catatan": "Perlengkapan kantor",
        },
        {
            "tanggal": date.today().isoformat(),
            "nama": "Kompresor Angin",
            "harga_beli": 5000000,
            "kategori": "PERALATAN",
            "catatan": "Peralatan bengkel",
        },
    ],
    "hutang_opening": [
        {
            "tanggal": date.today().isoformat(),
            "nama_kreditur": "Supplier X",
            "nominal": 1500000,
            "unit": "BENGKEL",
            "catatan": "Sisa hutang pembelian part",
        },
        {
            "tanggal": date.today().isoformat(),
            "nama_kreditur": "Dealer Mobil Y",
            "nominal": 50000000,
            "unit": "JUAL_BELI_MOBIL",
            "catatan": "Sisa hutang pembelian mobil",
        },
        {
            "tanggal": date.today().isoformat(),
            "nama_kreditur": "Investor Pak Z",
            "nominal": 10000000,
            "unit": "JUAL_BELI_MOBIL",
            "catatan": "Hutang dana investor",
        },
    ],
    "piutang_opening": [
        {
            "tanggal": date.today().isoformat(),
            "nama_debitur": "Pelanggan Y",
            "nominal": 750000,
            "unit": "BENGKEL",
            "catatan": "Sisa piutang servis",
        },
        {
            "tanggal": date.today().isoformat(),
            "nama_debitur": "Perusahaan Angkut A",
            "nominal": 5000000,
            "unit": "JASA_ANGKUT",
            "catatan": "Sisa piutang jasa angkut",
        },
        {
            "tanggal": date.today().isoformat(),
            "nama_debitur": "Pak Pembeli Mobil",
            "nominal": 20000000,
            "unit": "JUAL_BELI_MOBIL",
            "catatan": "Sisa piutang jual-beli mobil",
        },
        {
            "tanggal": date.today().isoformat(),
            "nama_debitur": "Karyawan Kasbon",
            "nominal": 500000,
            "unit": "KASBON",
            "catatan": "Kasbon karyawan",
        },
    ],
    "armada": [
        {"nama": "Truk 01", "nopol": "F 1234 XX", "jenis": "Dump Truck", "is_active": "ya"},
        {"nama": "Truk 02", "nopol": "F 5678 YY", "jenis": "Box", "is_active": "ya"},
    ],
    "supir": [
        {
            "kode": "SPR-DRV-01",
            "nama": "Joko Supir",
            "telepon": "0822222222",
            "tanggal_bergabung": date.today().isoformat(),
            "is_active": "ya",
        },
        {
            "kode": "SPR-DRV-02",
            "nama": "Bambang",
            "telepon": "0833333333",
            "tanggal_bergabung": date.today().isoformat(),
            "nopol_kendaraan": "F 5678 YY",
            "is_active": "ya",
        },
    ],
    "mobil": [
        {
            "kode": "MBL-001",
            "merek": "Toyota",
            "model": "Avanza",
            "tahun": 2018,
            "warna": "Hitam",
            "nomor_plat": "F 9999 AA",
            "harga_beli": 120000000,
            "harga_jual": 135000000,
            "tanggal_masuk": date.today().isoformat(),
            "status": "TERSEDIA",
            "tipe_kepemilikan": "TPM",
        },
        {
            "kode": "MBL-002",
            "merek": "Daihatsu",
            "model": "Xenia",
            "tahun": 2019,
            "warna": "Putih",
            "nomor_plat": "F 8888 BB",
            "harga_beli": 110000000,
            "harga_jual": 125000000,
            "tanggal_masuk": date.today().isoformat(),
            "status": "TERSEDIA",
            "tipe_kepemilikan": "INVESTOR",
            "nama_investor": "Pak Z",
            "nominal_investor": 100000000,
        },
        {
            "kode": "MBL-003",
            "merek": "Suzuki",
            "model": "Ertiga",
            "tahun": 2020,
            "warna": "Silver",
            "nomor_plat": "F 7777 CC",
            "harga_beli": 140000000,
            "harga_jual": 155000000,
            "tanggal_masuk": date.today().isoformat(),
            "status": "TERSEDIA",
            "tipe_kepemilikan": "TPM",
            "biaya pengeluaran/ operasional": 500000,
            "keterangan biaya pengeluaran/ operasional": "Biaya manajemen unit & administrasi",
            "biaya part dan service": 750000,
            "catatan": "Perbaikan di bengkel: ganti oli, filter, dan servis AC sebelum dijual",
        },
    ],
    "neraca_check": [
        {
            # AKTIVA LANCAR - Kas & Bank
            "kas_tunai": 50000000,
            "kas_bank": 10000000,
            "kas_unit_bengkel": 2500000,
            "kas_unit_jasa_angkut": 1500000,
            "kas_unit_mobil": 2500000,
            "total_kas_bank": 66500000,
            # AKTIVA LANCAR - Piutang
            "piutang_lainnya": 5000000,
            "piutang_karyawan": 3000000,
            "piutang_usaha": 8000000,
            "piutang_mobil": 5000000,
            "piutang_jasa_angkut": 5250000,
            "total_piutang": 26250000,
            # AKTIVA LANCAR - Persediaan & Stok
            "persediaan_sparepart": 75000000,
            "stok_mobil": 150000000,
            "stok_mobil_harga_beli": 130000000,
            "stok_mobil_biaya_persiapan": 10000000,
            "stok_mobil_perbaikan_external": 5000000,
            "stok_mobil_perbaikan_internal": 5000000,
            "total_aktiva_lancar": 317750000,
            # AKTIVA TETAP
            "total_aset_tetap": 100000000,
            "aset_detail_kode": "ASET-001",
            "aset_detail_nama": "Kompressor",
            "aset_detail_harga_beli": 100000000,
            "total_aktiva": 417750000,
            # HUTANG
            "hutang_part": 15000000,
            "hutang_mobil": 20000000,
            "hutang_investor": 10000000,
            "hutang_lainnya": 10000000,
            "hutang_jasa_angkut": 5000000,
            "uang_muka_penjualan": 0,
            "piutang_booking": 0,
            "total_hutang": 60000000,
            # MODAL
            "setoran_modal_kas": 200000000,
            "setoran_modal_non_kas": 87750000,
            "setoran_modal_total": 287750000,
            "laba_ditahan": 0,
            "prive": 0,
            "total_modal": 287750000,
            # TOTAL PASIVA & BALANCE CHECK
            "total_pasiva": 347750000,
            "selisih": 70000000,
            "catatan": "Verifikasi neraca opening balance - sesuaikan dengan kondisi nyata",
        },
    ],
}

INSTRUCTIONS = [
    "TEMPLATE IMPORT DATA EXISTING - TPM Super App",
    "",
    "CARA PAKAI:",
    "1. Isi sheet data (customers, suppliers, spare_parts, ...).",
    "2. Baris contoh (warna kuning) boleh dihapus / diganti / ditambah.",
    "3. Isi sheet neraca_check sebagai referensi verifikasi (opsional).",
    "4. Upload di Settings > Import Data > Preview (dry-run).",
    "5. Periksa hasil preview: cek error + cross-check neraca_check.",
    "6. Perbaiki error yang muncul, lalu Commit.",
    "",
    "URUTAN IMPOR (otomatis):",
    "customers > suppliers > spare_parts > jasa_servis > karyawan >",
    "kas_opening > hutang_opening > piutang_opening > armada > supir > mobil > neraca_check",
    "",
    "ATURAN PENTING:",
    "- MVP = master + stok + opening finance (bukan full histori transaksi).",
    "- kas_opening menambah MASUK ke dompet (idempotent per batch+jenis_kas).",
    "- hutang/piutang_opening mencatat sisa (total_dibayar=0, sisa=nominal).",
    "- spare_parts: upsert by kode (jika kosong > generate SPR...).",
    "- customers/suppliers: upsert by kode (jika kosong > generate).",
    "- tanggal format: YYYY-MM-DD.",
    "- Jangan ubah nama sheet.",
    "",
    "JENIS KAS AKTIF (kolom jenis_kas di kas_opening):",
    "- KAS_UTAMA      : kas tunai kantor utama",
    "- BANK_UTAMA     : rekening bank utama (BCA)",
    "- KAS_UNIT_BENGKEL      : laci kas unit bengkel",
    "- KAS_UNIT_JASA_ANGKUT  : kas unit jasa angkut",
    "- KAS_UNIT_MOBIL        : kas unit jual-beli mobil",
    "- CASH / BANK_BCA / BANK_MANDIRI / BANK_BRI / BANK_LAINNYA : dompet tambahan (opsional)",
    "",
    "UNIT hutang/piutang (kolom unit):",
    "- BENGKEL          : unit bengkel",
    "- JASA_ANGKUT      : unit jasa angkut",
    "- JUAL_BELI_MOBIL  : unit jual-beli mobil",
    "- LAINNYA          : hutang/piutang manual lainnya",
    "- KASBON           : (piutang saja) kasbon/pinjaman karyawan",
    "",
    "SHEET MOBIL - KOLOM BIAYA:",
    "- 'biaya pengeluaran/ operasional'  : biaya manajemen / biaya unit / operasional mobil",
    "- 'keterangan biaya pengeluaran/ operasional' : keterangan biaya manajemen tsb",
    "- 'biaya part dan service'          : biaya perbaikan part & servis di bengkel",
    "- Kolom biaya di atas hanya dicatat ke catatan mobil (opsional).",
    "  Untuk biaya perbaikan bengkel & biaya manajemen unit yang lengkap,",
    "  gunakan menu transaksi bengkel / biaya unit setelah import.",
    "",
    "KESEIMBANGAN NERACA:",
    "Agar laporan Neraca & Perubahan Modal seimbang, pastikan kombinasi",
    "kas_opening + asset + hutang_opening + piutang_opening + stok mobil",
    "mencerminkan kondisi keuangan nyata. Umumnya:",
    "  Total Aktiva = Kas + Piutang + Stok + Aset Tetap",
    "  Total Pasiva = Hutang + Modal (Setoran Kas + Non-Kas)",
    "Backend otomatis menyeimbangkan modal non-kas dari aset opening yang",
    "diimport tanpa entry kas (piutang/hutang/aset tetap).",
    "",
    "SHEET NERACA_CHECK (opsional):",
    "Isi sheet ini sebagai referensi verifikasi setelah import. Struktur mengikuti neraca.tsx.",
    "",
    "Kolom yang tersedia:",
    "  AKTIVA LANCAR: kas_tunai, kas_bank, kas_unit_bengkel, kas_unit_jasa_angkut,",
    "    kas_unit_mobil, total_kas_bank,",
    "    piutang_lainnya, piutang_karyawan, piutang_usaha, piutang_mobil,",
    "    piutang_jasa_angkut, total_piutang,",
    "    persediaan_sparepart, stok_mobil (+ breakdown: harga_beli, biaya_persiapan,",
    "    perbaikan_external, perbaikan_internal), total_aktiva_lancar",
    "  AKTIVA TETAP: total_aset_tetap, aset_detail_kode/nama/harga_beli",
    "  HUTANG: hutang_part, hutang_mobil, hutang_investor, hutang_lainnya,",
    "    hutang_jasa_angkut, uang_muka_penjualan, piutang_booking, total_hutang",
    "  MODAL: setoran_modal_kas, setoran_modal_non_kas, setoran_modal_total,",
    "    laba_ditahan, prive, total_modal",
    "  BALANCE CHECK: total_pasiva, selisih",
    "",
    "Backend otomatis menghitung (auto-sum) dari sheet kas_opening, piutang_opening,",
    "hutang_opening, asset, mobil. Nilai di neraca_check digunakan sebagai ekspektasi",
    "untuk cross-check. Jika ada selisih > Rp 100, muncul warning di preview.",
    "Sheet ini TIDAK insert data ke database, hanya untuk verifikasi.",
    "",
    "ADMIN ONLY. Backup database sebelum commit production.",
]


class DataImportService:
    def __init__(self, db: Session):
        self.db = db

    # ------------------------------------------------------------------ utils
    @staticmethod
    def _cell(v: Any) -> Optional[str]:
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    @staticmethod
    def _dec(v: Any, default: str = "0") -> Decimal:
        if v is None or str(v).strip() == "":
            return Decimal(default)
        try:
            return Decimal(str(v).strip().replace(",", "."))
        except (InvalidOperation, ValueError):
            raise ValueError(f"angka tidak valid: {v}")

    @staticmethod
    def _int(v: Any, default: int = 0) -> int:
        if v is None or str(v).strip() == "":
            return default
        return int(Decimal(str(v).strip().replace(",", ".")))

    @staticmethod
    def _date(v: Any, default: Optional[date] = None) -> date:
        if v is None or str(v).strip() == "":
            return default or date.today()
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        s = str(v).strip()[:10]
        return datetime.strptime(s, "%Y-%m-%d").date()

    @staticmethod
    def _bool(v: Any, default: bool = True) -> bool:
        if v is None or str(v).strip() == "":
            return default
        return str(v).strip().lower() in ("1", "true", "ya", "yes", "y", "aktif", "v")

    def _next_kode(self, prefix: str, model, field: str = "kode") -> str:
        date_str = datetime.now().strftime("%y%m")
        like = f"{prefix}{date_str}%"
        col = getattr(model, field)
        last = (
            self.db.query(model)
            .filter(col.like(like))
            .order_by(col.desc())
            .first()
        )
        n = 1
        if last:
            try:
                n = int(str(getattr(last, field))[-4:]) + 1
            except Exception:
                n = 1
        return f"{prefix}{date_str}{n:04d}"

    def _gen_nomor(self, prefix_key: str, model, field: str) -> str:
        today = datetime.now()
        prefix = TRANSACTION_PREFIXES[prefix_key]
        date_str = today.strftime("%y%m%d")
        col = getattr(model, field)
        last = (
            self.db.query(model)
            .filter(col.like(f"{prefix}{date_str}%"))
            .order_by(model.id.desc())
            .first()
        )
        n = 1
        if last:
            try:
                n = int(str(getattr(last, field))[-4:]) + 1
            except Exception:
                n = 1
        return f"{prefix}{date_str}{n:04d}"

    def _map_header(self, sheet) -> Dict[str, int]:
        header = [str(c.value or "").strip().lower() for c in sheet[1]]
        return {h: i for i, h in enumerate(header) if h}

    def _row_dict(self, row, header_map: Dict[str, int], fields: List[Tuple[str, str, bool]]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        for header, key, _req in fields:
            idx = header_map.get(header.lower())
            if idx is None or idx >= len(row):
                out[key] = None
            else:
                out[key] = row[idx]
        return out

    def _empty_row(self, row) -> bool:
        return not any(c is not None and str(c).strip() != "" for c in row)

    # ------------------------------------------------ neraca formula helper
    @staticmethod
    def _write_neraca_formulas(ws, fields):
        """Write Excel formulas in row 2 of neraca_check sheet for auto-sum.
        
        Structure matches neraca.tsx exactly.
        Column layout of source sheets:
          kas_opening:    A=tanggal, B=jenis_kas, C=nominal, D=keterangan, E=catatan
          piutang_opening: A=tanggal, B=nama_debitur, C=nominal, D=unit, E=telepon...
          hutang_opening:  A=tanggal, B=nama_kreditur, C=nominal, D=unit, E=telepon...
          asset:           A=kode, B=tanggal, C=nama, D=nominal(harga_beli), E=kategori
          mobil:           A=kode, B=merek, C=model, D=tahun, E=warna, F=plat, G=mesin, H=raga, I=harga_beli
        """
        field_col = {}
        for col, (header, key, _r) in enumerate(fields, start=1):
            field_col[key] = col

        def _set_formula(header_text, formula):
            col = field_col.get(header_text)
            if col:
                ws.cell(2, col, formula)
                ws.cell(2, col).fill = PatternFill("solid", fgColor="D1FAE5")
                ws.cell(2, col).number_format = '#,##0'

        # === Kas & Bank ===
        _set_formula('kas_tunai', '=SUMIF(kas_opening!B:B,"KAS_UTAMA",kas_opening!C:C)+SUMIF(kas_opening!B:B,"CASH",kas_opening!C:C)')
        _set_formula('kas_bank', '=SUMIF(kas_opening!B:B,"BANK_UTAMA",kas_opening!C:C)')
        _set_formula('unit_cash', '=SUMIF(kas_opening!B:B,"KAS_UNIT_BENGKEL",kas_opening!C:C)+SUMIF(kas_opening!B:B,"KAS_UNIT_JASA_ANGKUT",kas_opening!C:C)+SUMIF(kas_opening!B:B,"KAS_UNIT_MOBIL",kas_opening!C:C)')
        _set_formula('total_kas_bank', '=A2+B2+C2')

        # === Piutang Usaha ===
        _set_formula('piutang_lainnya', '=SUMIF(piutang_opening!D:D,"LAINNYA",piutang_opening!C:C)')
        _set_formula('piutang_karyawan', '=SUMIF(piutang_opening!D:D,"KASBON",piutang_opening!C:C)')
        _set_formula('piutang_usaha', '=SUMIF(piutang_opening!D:D,"BENGKEL",piutang_opening!C:C)')
        _set_formula('piutang_mobil', '=SUMIF(piutang_opening!D:D,"JUAL_BELI_MOBIL",piutang_opening!C:C)')
        _set_formula('piutang_jasa_angkut', '=SUMIF(piutang_opening!D:D,"JASA_ANGKUT",piutang_opening!C:C)')
        _set_formula('total_piutang', '=E2+F2+G2+H2+I2')

        # === Persediaan & Stok ===
        # spare_parts: G=stok, I=harga_beli (row 2+ to skip header text)
        _set_formula('persediaan_sparepart', '=SUMPRODUCT(spare_parts!G2:G10000,spare_parts!I2:I10000)')
        # mobil: I=harga_beli
        _set_formula('stok_mobil', '=SUM(mobil!I2:I10000)')
        _set_formula('stok_mobil_detail_harga_beli', '=SUM(mobil!I2:I10000)')
        _set_formula('stok_mobil_detail_biaya_persiapan', '0')
        _set_formula('stok_mobil_detail_perbaikan_external', '0')
        _set_formula('stok_mobil_detail_perbaikan_internal', '0')
        _set_formula('total_aktiva_lancar', '=D2+J2+K2+L2')

        # === Aktiva Tetap ===
        _set_formula('detail_aset_kode', '')
        _set_formula('detail_aset_nama', '')
        _set_formula('detail_aset_harga_beli', '')
        _set_formula('total_aset_tetap', '=SUM(asset!D:D)')
        _set_formula('total_aktiva', '=Q2+U2')

        # === Modal (konsep saldo awal: Modal = Total Aktiva - Total Hutang) ===
        _set_formula('setoran_modal', '0')
        _set_formula('setoran_modal_kas', '0')
        _set_formula('modal_non_kas', '0')
        _set_formula('modal_persediaan', '0')
        _set_formula('modal_stok_mobil', '0')
        _set_formula('modal_aset_tetap', '0')
        _set_formula('laba_ditahan', '0')
        _set_formula('prive', '0')
        _set_formula('total_modal', '=V2-AM2')

        # === Hutang ===
        _set_formula('hutang_part', '=SUMIF(hutang_opening!D:D,"BENGKEL",hutang_opening!C:C)')
        _set_formula('hutang_mobil', '=SUMIF(hutang_opening!D:D,"JUAL_BELI_MOBIL",hutang_opening!C:C)')
        _set_formula('hutang_investor', '=SUMIF(hutang_opening!D:D,"INVESTOR",hutang_opening!C:C)')
        _set_formula('hutang_lainnya', '=SUMIF(hutang_opening!D:D,"LAINNYA",hutang_opening!C:C)')
        _set_formula('hutang_jasa_angkut', '=SUMIF(hutang_opening!D:D,"JASA_ANGKUT",hutang_opening!C:C)')
        _set_formula('uang_muka_penjualan', '0')
        _set_formula('piutang_booking', '0')
        _set_formula('total_hutang', '=AF2+AG2+AH2+AI2+AJ2+AK2+AL2')

        # === Balance Check ===
        _set_formula('total_pasiva', '=AM2+AE2')
        _set_formula('selisih', '=V2-AN2')

    # ----------------------------------------------------------- template
    def generate_template(self) -> io.BytesIO:
        wb = openpyxl.Workbook()
        # instruction sheet
        ws0 = wb.active
        ws0.title = "_INSTRUKSI"
        ws0["A1"] = "\n".join(INSTRUCTIONS)
        ws0["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws0.column_dimensions["A"].width = 100
        ws0.row_dimensions[1].height = 320

        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_font = Font(bold=True, color="FFFFFF")
        example_fill = PatternFill("solid", fgColor="FEF3C7")

        for sheet_name in SHEET_ORDER:
            if sheet_name == "_INSTRUKSI":
                continue
            ws = wb.create_sheet(sheet_name)
            fields = SHEET_HEADERS[sheet_name]
            for col, (header, _k, _r) in enumerate(fields, start=1):
                cell = ws.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[get_column_letter(col)].width = max(14, len(header) + 2)
            for r_i, example in enumerate(EXAMPLE_ROWS.get(sheet_name, []), start=2):
                for c_i, (header, key, _r) in enumerate(fields, start=1):
                    cell = ws.cell(r_i, c_i, example.get(key))
                    cell.fill = example_fill
            # neraca_check: overwrite example row with Excel auto-sum formulas
            if sheet_name == "neraca_check":
                DataImportService._write_neraca_formulas(ws, fields)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ----------------------------------------------------------- parse workbook
    def _parse_workbook(self, file_content: bytes) -> Tuple[Dict[str, List[Dict[str, Any]]], List[str]]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Gagal membaca Excel: {e}",
            )
        parsed: Dict[str, List[Dict[str, Any]]] = {}
        known_sheets = {s for s in SHEET_ORDER if s != "_INSTRUKSI"}
        unknown_sheets: List[str] = []
        for sheet_name in wb.sheetnames:
            if sheet_name not in known_sheets:
                unknown_sheets.append(sheet_name)
                continue
            sheet = wb[sheet_name]
            fields = SHEET_HEADERS[sheet_name]
            header_map = self._map_header(sheet)
            # require at least required headers
            missing_headers = [h for h, k, req in fields if req and h.lower() not in header_map]
            if missing_headers:
                raise HTTPException(
                    status_code=400,
                    detail=f"Sheet '{sheet_name}' kurang kolom: {', '.join(missing_headers)}",
                )
            rows: List[Dict[str, Any]] = []
            start_row = 3 if sheet_name == "neraca_check" else 2
            for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if self._empty_row(row):
                    continue
                d = self._row_dict(row, header_map, fields)
                d["_row"] = row_idx
                rows.append(d)
            parsed[sheet_name] = rows
        return parsed, unknown_sheets

    # ----------------------------------------------------------- validate row
    def _validate_row(self, sheet: str, row: Dict[str, Any]) -> Optional[str]:
        fields = SHEET_HEADERS[sheet]
        r = row["_row"]
        for header, key, req in fields:
            if req and (row.get(key) is None or str(row.get(key)).strip() == ""):
                return f"Baris {r}: kolom '{header}' wajib diisi"
        try:
            if sheet == "spare_parts":
                self._dec(row.get("stok"), "0")
                self._dec(row.get("harga_beli"), "0")
                self._dec(row.get("harga_jual"), "0")
            elif sheet == "jasa_servis":
                self._dec(row.get("harga"), "0")
            elif sheet == "karyawan":
                self._dec(row.get("gaji_pokok"), "0")
                self._date(row.get("tanggal_bergabung"), date.today())
                st = self._cell(row.get("status")) or "AKTIF"
                if st.upper().replace(" ", "_") not in {e.value for e in EmployeeStatus}:
                    return f"Baris {r}: status karyawan tidak valid ({st})"
            elif sheet == "kas_opening":
                self._date(row.get("tanggal"))
                jenis = (self._cell(row.get("jenis_kas")) or "").upper().replace(" ", "_")
                if jenis not in {e.value for e in KasBankJenis}:
                    return f"Baris {r}: jenis_kas tidak valid ({jenis})"
                nom = self._dec(row.get("nominal"))
                if nom <= 0:
                    return f"Baris {r}: nominal harus > 0"
            elif sheet == "asset":
                self._date(row.get("tanggal"), date.today())
                nom = self._dec(row.get("harga_beli"))
                if nom <= 0:
                    return f"Baris {r}: nominal harus > 0"
                kat = (self._cell(row.get("kategori")) or "LAINNYA").upper().replace(" ", "_")
                # alias user-friendly names â†’ enum values
                KAT_ALIASES = {"ELEKTRONIK": "ELECTRONIC"}
                kat = KAT_ALIASES.get(kat, kat)
                if kat not in {e.value for e in AssetCategory}:
                    return f"Baris {r}: kategori tidak valid ({kat})"
            elif sheet in ("hutang_opening", "piutang_opening"):
                self._date(row.get("tanggal"))
                nom = self._dec(row.get("nominal"))
                if nom <= 0:
                    return f"Baris {r}: nominal harus > 0"
                unit = (self._cell(row.get("unit")) or "BENGKEL").upper().replace(" ", "_")
                # map aliases
                UNIT_ALIASES = {
                    "MOBIL": "JUAL_BELI_MOBIL",
                    "INVESTOR": "MODAL",
                    "KARYAWAN": "KASBON",
                    "PART_JB_MOBIL": "PEMBELIAN_PART",
                }
                unit = UNIT_ALIASES.get(unit, unit)
                if unit not in {e.value for e in KasBankSource}:
                    return f"Baris {r}: unit tidak valid ({unit})"
            elif sheet == "armada":
                if not self._cell(row.get("nopol")):
                    return f"Baris {r}: nopol wajib"
            elif sheet == "supir":
                self._date(row.get("tanggal_bergabung"), date.today())
            elif sheet == "mobil":
                self._int(row.get("tahun"), 0)
                self._dec(row.get("harga_beli"))
                st = (self._cell(row.get("status")) or "TERSEDIA").upper()
                if st not in {e.value for e in CarStatus}:
                    return f"Baris {r}: status mobil tidak valid ({st})"
                self._date(row.get("tanggal_masuk"), date.today())
        except Exception as e:
            return f"Baris {r}: {e}"
        return None

    # ----------------------------------------------------------- apply rows
    def _apply_customers(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("customers", row)
            if err:
                res["errors"].append(err)
                continue
            nama = self._cell(row["nama"])
            kode = self._cell(row.get("kode"))
            tipe = (self._cell(row.get("tipe")) or "perorangan").lower()
            if tipe not in ("perorangan", "perusahaan"):
                tipe = "perorangan"
            existing = None
            if kode:
                existing = self.db.query(Customer).filter(Customer.kode == kode).first()
            if not existing and not kode:
                # try match by nama exact
                existing = self.db.query(Customer).filter(Customer.nama == nama).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            if existing:
                existing.nama = nama
                existing.tipe = tipe
                existing.telepon = self._cell(row.get("telepon"))
                existing.alamat = self._cell(row.get("alamat"))
                existing.kota = self._cell(row.get("kota"))
                existing.email = self._cell(row.get("email"))
                existing.npwp = self._cell(row.get("npwp"))
                existing.catatan = self._cell(row.get("catatan"))
                res["updated"] += 1
            else:
                if not kode:
                    kode = self._next_kode("CUS", Customer)
                self.db.add(
                    Customer(
                        kode=kode,
                        nama=nama,
                        tipe=tipe,
                        telepon=self._cell(row.get("telepon")),
                        alamat=self._cell(row.get("alamat")),
                        kota=self._cell(row.get("kota")),
                        email=self._cell(row.get("email")),
                        npwp=self._cell(row.get("npwp")),
                        catatan=self._cell(row.get("catatan")),
                    )
                )
                res["created"] += 1
        return res

    def _apply_suppliers(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("suppliers", row)
            if err:
                res["errors"].append(err)
                continue
            nama = self._cell(row["nama"])
            kode = self._cell(row.get("kode"))
            existing = None
            if kode:
                existing = self.db.query(Supplier).filter(Supplier.kode == kode).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            if existing:
                existing.nama = nama
                existing.telepon = self._cell(row.get("telepon"))
                existing.alamat = self._cell(row.get("alamat"))
                existing.kota = self._cell(row.get("kota"))
                existing.email = self._cell(row.get("email"))
                existing.contact_person = self._cell(row.get("contact_person"))
                existing.bank = self._cell(row.get("bank"))
                existing.rekening = self._cell(row.get("rekening"))
                existing.catatan = self._cell(row.get("catatan"))
                res["updated"] += 1
            else:
                if not kode:
                    kode = self._next_kode("SUP", Supplier)
                self.db.add(
                    Supplier(
                        kode=kode,
                        nama=nama,
                        telepon=self._cell(row.get("telepon")),
                        alamat=self._cell(row.get("alamat")),
                        kota=self._cell(row.get("kota")),
                        email=self._cell(row.get("email")),
                        contact_person=self._cell(row.get("contact_person")),
                        bank=self._cell(row.get("bank")),
                        rekening=self._cell(row.get("rekening")),
                        catatan=self._cell(row.get("catatan")),
                    )
                )
                res["created"] += 1
        return res

    def _apply_spare_parts(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("spare_parts", row)
            if err:
                res["errors"].append(err)
                continue
            try:
                nama = self._cell(row["nama"])
                kode = self._cell(row.get("kode"))
                stok = self._dec(row.get("stok"), "0")
                stok_min = self._int(row.get("stok_minimum"), 5)
                hb = self._dec(row.get("harga_beli"), "0")
                hj = self._dec(row.get("harga_jual"), "0")
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            existing = None
            if kode:
                existing = self.db.query(SparePart).filter(SparePart.kode == kode).first()
            if not existing:
                kp = self._cell(row.get("kode_part"))
                if kp:
                    existing = self.db.query(SparePart).filter(SparePart.kode_part == kp).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            if existing:
                existing.nama = nama
                existing.kode_part = self._cell(row.get("kode_part"))
                existing.kategori = self._cell(row.get("kategori")) or "Umum"
                existing.merek = self._cell(row.get("merek"))
                existing.satuan = self._cell(row.get("satuan")) or "pcs"
                existing.stok = stok
                existing.stok_minimum = stok_min
                existing.harga_beli = hb
                existing.harga_jual = hj
                existing.lokasi_rak = self._cell(row.get("lokasi_rak"))
                existing.catatan = self._cell(row.get("catatan"))
                res["updated"] += 1
            else:
                if not kode:
                    kode = self._next_kode("SPR", SparePart)
                self.db.add(
                    SparePart(
                        kode=kode,
                        nama=nama,
                        kode_part=self._cell(row.get("kode_part")),
                        kategori=self._cell(row.get("kategori")) or "Umum",
                        merek=self._cell(row.get("merek")),
                        satuan=self._cell(row.get("satuan")) or "pcs",
                        stok=stok,
                        stok_minimum=stok_min,
                        harga_beli=hb,
                        harga_jual=hj,
                        lokasi_rak=self._cell(row.get("lokasi_rak")),
                        catatan=self._cell(row.get("catatan")),
                    )
                )
                res["created"] += 1
        return res

    def _apply_jasa_servis(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("jasa_servis", row)
            if err:
                res["errors"].append(err)
                continue
            nama = self._cell(row["nama"])
            try:
                harga = self._dec(row.get("harga"), "0")
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            existing = self.db.query(JasaServis).filter(JasaServis.nama == nama).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            if existing:
                existing.kategori = self._cell(row.get("kategori"))
                existing.harga = harga
                existing.deskripsi = self._cell(row.get("deskripsi"))
                res["updated"] += 1
            else:
                self.db.add(
                    JasaServis(
                        nama=nama,
                        kategori=self._cell(row.get("kategori")),
                        harga=harga,
                        deskripsi=self._cell(row.get("deskripsi")),
                    )
                )
                res["created"] += 1
        return res

    def _apply_karyawan(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("karyawan", row)
            if err:
                res["errors"].append(err)
                continue
            nama = self._cell(row["nama"])
            jabatan = self._cell(row["jabatan"]) or "Staff"
            kode = self._cell(row.get("kode"))
            existing = None
            if kode:
                existing = self.db.query(Karyawan).filter(Karyawan.kode == kode).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            try:
                tgl = self._date(row.get("tanggal_bergabung"), date.today())
                gaji = self._dec(row.get("gaji_pokok"), "0")
                tunj = self._dec(row.get("tunjangan"), "0")
                st = (self._cell(row.get("status")) or "AKTIF").upper().replace(" ", "_")
                status = EmployeeStatus(st)
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            if existing:
                existing.nama = nama
                existing.jabatan = jabatan
                existing.nik = self._cell(row.get("nik"))
                existing.telepon = self._cell(row.get("telepon"))
                existing.alamat = self._cell(row.get("alamat"))
                existing.email = self._cell(row.get("email"))
                existing.gaji_pokok = gaji
                existing.tunjangan = tunj
                existing.status = status
                existing.catatan = self._cell(row.get("catatan"))
                res["updated"] += 1
            else:
                if not kode:
                    kode = self._next_kode("KRY", Karyawan)
                self.db.add(
                    Karyawan(
                        kode=kode,
                        nama=nama,
                        jabatan=jabatan,
                        nik=self._cell(row.get("nik")),
                        telepon=self._cell(row.get("telepon")),
                        alamat=self._cell(row.get("alamat")),
                        email=self._cell(row.get("email")),
                        tanggal_bergabung=tgl,
                        gaji_pokok=gaji,
                        tunjangan=tunj,
                        status=status,
                        catatan=self._cell(row.get("catatan")),
                    )
                )
                res["created"] += 1
        return res

    def _apply_kas_opening(
        self, rows: List[Dict[str, Any]], dry: bool, batch_id: str, user_id: Optional[int]
    ) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("kas_opening", row)
            if err:
                res["errors"].append(err)
                continue
            try:
                tgl = self._date(row.get("tanggal"))
                jenis = KasBankJenis((self._cell(row.get("jenis_kas")) or "").upper().replace(" ", "_"))
                nominal = self._dec(row.get("nominal"))
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            ref = f"IMP-{batch_id}-KAS-{jenis.value}"
            existing = (
                self.db.query(KasBank)
                .filter(KasBank.nomor_referensi == ref)
                .first()
            )
            if existing:
                res["skipped"] += 1
                continue
            if dry:
                res["created"] += 1
                continue
            saldo_sebelum = KasBank.get_current_balance(self.db, jenis)
            nomor = self._gen_nomor("kas_bank", KasBank, "nomor_transaksi")
            ket = self._cell(row.get("keterangan")) or f"Saldo opening import {jenis.value}"
            kb = KasBank(
                nomor_transaksi=nomor,
                tanggal=tgl,
                jenis=jenis,
                metode_bayar=PaymentMethod.TUNAI,
                tipe=KasBankType.MASUK,
                nominal=nominal,
                sumber=KasBankSource.MODAL,
                nomor_referensi=ref,
                keterangan=ket,
                catatan=self._cell(row.get("catatan")) or f"import_batch={batch_id}",
                created_by=user_id,
            )
            kb.calculate_saldo(saldo_sebelum)
            self.db.add(kb)
            self.db.flush()
            res["created"] += 1
        return res

    def _apply_aset(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("asset", row)
            if err:
                res["errors"].append(err)
                continue
            try:
                tgl = self._date(row.get("tanggal"), date.today())
                nama = self._cell(row["nama"])
                hb = self._dec(row.get("harga_beli"))
                kat_s = (self._cell(row.get("kategori")) or "LAINNYA").upper().replace(" ", "_")
                KAT_ALIASES = {"ELEKTRONIK": "ELECTRONIC"}
                kat_s = KAT_ALIASES.get(kat_s, kat_s)
                kat = AssetCategory(kat_s)
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            existing = None
            kode = self._cell(row.get("kode"))
            if kode:
                existing = self.db.query(Aset).filter(Aset.kode == kode).first()
            if not existing:
                existing = self.db.query(Aset).filter(Aset.nama == nama).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            if existing:
                existing.tanggal_beli = tgl
                existing.harga_beli = hb
                existing.kategori = kat
                existing.catatan = self._cell(row.get("catatan"))
                res["updated"] += 1
            else:
                if not kode:
                    # Cari kode unik yang belum ada di database
                    date_str = datetime.now().strftime("%y%m")
                    like = f"AST{date_str}%"
                    last = (
                        self.db.query(Aset)
                        .filter(Aset.kode.like(like))
                        .order_by(Aset.kode.desc())
                        .first()
                    )
                    n = 1
                    if last:
                        try:
                            n = int(str(last.kode)[-4:]) + 1
                        except Exception:
                            n = 1
                    # Double check loop biar aman dari tabrakan
                    while True:
                        kode = f"AST{date_str}{n:04d}"
                        if not self.db.query(Aset).filter(Aset.kode == kode).first():
                            break
                        n += 1

                self.db.add(
                    Aset(
                        kode=kode,
                        nama=nama,
                        tanggal_beli=tgl,
                        harga_beli=hb,
                        kategori=kat,
                        umur_ekonomis=4,
                        nilai_residu=Decimal("0"),
                        status=AssetStatus.AKTIF,
                        catatan=self._cell(row.get("catatan")),
                    )
                )
                self.db.flush()
                res["created"] += 1
        return res

    def _map_unit_alias(self, unit_str: str) -> str:
        """Map user-friendly unit names to valid KasBankSource enum values."""
        UNIT_ALIASES = {
            "MOBIL": "JUAL_BELI_MOBIL",
            "INVESTOR": "MODAL",
            "KARYAWAN": "KASBON",
            "PART_JB_MOBIL": "PEMBELIAN_PART",
        }
        return UNIT_ALIASES.get(unit_str, unit_str)

    def _apply_hutang_opening(
        self, rows: List[Dict[str, Any]], dry: bool, batch_id: str, user_id: Optional[int]
    ) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("hutang_opening", row)
            if err:
                res["errors"].append(err)
                continue
            try:
                tgl = self._date(row.get("tanggal"))
                nama = self._cell(row["nama_kreditur"])
                nominal = self._dec(row.get("nominal"))
                unit_s = (self._cell(row.get("unit")) or "BENGKEL").upper().replace(" ", "_")
                unit_s = self._map_unit_alias(unit_s)
                unit = KasBankSource(unit_s)
                jth = None
                if row.get("tanggal_jatuh_tempo") not in (None, ""):
                    jth = self._date(row.get("tanggal_jatuh_tempo"))
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            ref = f"IMP-{batch_id}-HTG-{row['_row']}-{re.sub(r'[^A-Z0-9]', '', nama.upper())[:10]}"
            # Pastikan panjang referensi maksimum 50 karakter
            ref = ref[:50]
            existing = (
                self.db.query(HutangUsaha)
                .filter(HutangUsaha.nomor_referensi == ref)
                .first()
            )
            if existing:
                res["skipped"] += 1
                continue
            if dry:
                res["created"] += 1
                continue
            nomor = self._gen_nomor("hutang", HutangUsaha, "nomor_hutang")
            self.db.add(
                HutangUsaha(
                    nomor_hutang=nomor,
                    tanggal=tgl,
                    sumber=HutangSource.LAINNYA,
                    nomor_referensi=ref,
                    unit=unit,
                    nama_kreditur=nama,
                    telepon_kreditur=self._cell(row.get("telepon")),
                    alamat_kreditur=self._cell(row.get("alamat")),
                    nominal_hutang=nominal,
                    total_dibayar=Decimal("0"),
                    sisa_hutang=nominal,
                    tanggal_jatuh_tempo=jth,
                    status=HutangStatus.BELUM_LUNAS,
                    catatan=(self._cell(row.get("catatan")) or "") + f" | import_batch={batch_id}",
                    created_by=user_id,
                )
            )
            self.db.flush()
            res["created"] += 1
        return res

    def _apply_piutang_opening(
        self, rows: List[Dict[str, Any]], dry: bool, batch_id: str, user_id: Optional[int]
    ) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("piutang_opening", row)
            if err:
                res["errors"].append(err)
                continue
            try:
                tgl = self._date(row.get("tanggal"))
                nama = self._cell(row["nama_debitur"])
                nominal = self._dec(row.get("nominal"))
                unit_s = (self._cell(row.get("unit")) or "BENGKEL").upper().replace(" ", "_")
                unit_s = self._map_unit_alias(unit_s)
                unit = KasBankSource(unit_s)
                jth = None
                if row.get("tanggal_jatuh_tempo") not in (None, ""):
                    jth = self._date(row.get("tanggal_jatuh_tempo"))
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            ref = f"IMP-{batch_id}-PTG-{row['_row']}-{re.sub(r'[^A-Z0-9]', '', nama.upper())[:10]}"
            # Pastikan panjang referensi maksimum 50 karakter
            ref = ref[:50]
            existing = (
                self.db.query(PiutangUsaha)
                .filter(PiutangUsaha.nomor_referensi == ref)
                .first()
            )
            if existing:
                res["skipped"] += 1
                continue
            if dry:
                res["created"] += 1
                continue
            nomor = self._gen_nomor("piutang", PiutangUsaha, "nomor_piutang")
            # Route imported piutang to its neraca line by unit, matching the
            # template's neraca_check mapping (KASBON/LAINNYA stay separate).
            if unit == KasBankSource.JASA_ANGKUT:
                sumber = PiutangSource.JASA_ANGKUT
            elif unit == KasBankSource.JUAL_BELI_MOBIL:
                sumber = PiutangSource.JUAL_BELI_MOBIL
            elif unit == KasBankSource.KASBON:
                sumber = PiutangSource.KASBON_KARYAWAN
            elif unit == KasBankSource.BENGKEL:
                sumber = PiutangSource.BENGKEL
            else:
                sumber = PiutangSource.LAINNYA
            self.db.add(
                PiutangUsaha(
                    nomor_piutang=nomor,
                    tanggal=tgl,
                    sumber=sumber,
                    nomor_referensi=ref,
                    unit=unit,
                    nama_debitur=nama,
                    telepon_debitur=self._cell(row.get("telepon")),
                    alamat_debitur=self._cell(row.get("alamat")),
                    nominal_piutang=nominal,
                    total_dibayar=Decimal("0"),
                    sisa_piutang=nominal,
                    tanggal_jatuh_tempo=jth,
                    status=PiutangStatus.BELUM_LUNAS,
                    catatan=(self._cell(row.get("catatan")) or "") + f" | import_batch={batch_id}",
                    created_by=user_id,
                )
            )
            self.db.flush()
            res["created"] += 1
        return res

    def _apply_armada(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("armada", row)
            if err:
                res["errors"].append(err)
                continue
            nopol = self._cell(row["nopol"])
            nama = self._cell(row["nama"])
            existing = self.db.query(ArmadaJasaAngkut).filter(ArmadaJasaAngkut.nopol == nopol).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            if existing:
                existing.nama = nama
                existing.jenis = self._cell(row.get("jenis"))
                existing.is_active = self._bool(row.get("is_active"), True)
                existing.catatan = self._cell(row.get("catatan"))
                res["updated"] += 1
            else:
                self.db.add(
                    ArmadaJasaAngkut(
                        nama=nama,
                        nopol=nopol,
                        jenis=self._cell(row.get("jenis")),
                        is_active=self._bool(row.get("is_active"), True),
                        catatan=self._cell(row.get("catatan")),
                    )
                )
                res["created"] += 1
        return res

    def _apply_supir(self, rows: List[Dict[str, Any]], dry: bool) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            err = self._validate_row("supir", row)
            if err:
                res["errors"].append(err)
                continue
            nama = self._cell(row["nama"])
            kode = self._cell(row.get("kode"))
            existing = None
            if kode:
                existing = self.db.query(Supir).filter(Supir.kode == kode).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            try:
                tgl = self._date(row.get("tanggal_bergabung"), date.today())
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            if existing:
                existing.nama = nama
                existing.telepon = self._cell(row.get("telepon"))
                existing.nik = self._cell(row.get("nik"))
                existing.nomor_sim = self._cell(row.get("nomor_sim"))
                existing.jenis_sim = self._cell(row.get("jenis_sim"))
                existing.nopol_kendaraan = self._cell(row.get("nopol_kendaraan"))
                existing.is_active = self._bool(row.get("is_active"), True)
                existing.catatan = self._cell(row.get("catatan"))
                res["updated"] += 1
            else:
                if not kode:
                    kode = self._next_kode("DRV", Supir)
                self.db.add(
                    Supir(
                        kode=kode,
                        nama=nama,
                        telepon=self._cell(row.get("telepon")),
                        nik=self._cell(row.get("nik")),
                        nomor_sim=self._cell(row.get("nomor_sim")),
                        jenis_sim=self._cell(row.get("jenis_sim")),
                        tanggal_bergabung=tgl,
                        nopol_kendaraan=self._cell(row.get("nopol_kendaraan")),
                        is_active=self._bool(row.get("is_active"), True),
                        catatan=self._cell(row.get("catatan")),
                    )
                )
                res["created"] += 1
        return res

    def _apply_mobil(
        self, rows: List[Dict[str, Any]], dry: bool, user_id: Optional[int]
    ) -> Dict[str, Any]:
        res = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        for row in rows:
            # Skip continuation/biaya rows (all required fields empty)
            required_keys = ["merek", "model", "tahun", "warna", "nomor_plat", "harga_beli"]
            if all(row.get(k) in (None, "") for k in required_keys):
                res["skipped"] += 1
                continue
            err = self._validate_row("mobil", row)
            if err:
                res["errors"].append(err)
                continue
            try:
                plat = self._cell(row["nomor_plat"])
                merek = self._cell(row["merek"])
                model = self._cell(row["model"])
                tahun = self._int(row.get("tahun"), date.today().year)
                warna = self._cell(row.get("warna")) or "-"
                hb = self._dec(row.get("harga_beli"))
                hj = self._dec(row.get("harga_jual"), "0") if row.get("harga_jual") not in (None, "") else None
                tgl = self._date(row.get("tanggal_masuk"), date.today())
                st = CarStatus((self._cell(row.get("status")) or "TERSEDIA").upper())
                own = OwnershipType((self._cell(row.get("tipe_kepemilikan")) or "TPM").upper())
                kode = self._cell(row.get("kode"))
                no_mesin = self._cell(row.get("nomor_mesin"))
                no_rangka = self._cell(row.get("nomor_rangka"))
                ni = self._dec(row.get("nominal_investor"), "0") if row.get("nominal_investor") not in (None, "") else Decimal("0")
                pi = self._dec(row.get("persentase_investor"), "0") if row.get("persentase_investor") not in (None, "") else Decimal("0")
                # Collect biaya notes into catatan (full biaya mgmt via dedicated form)
                biaya_parts = []
                if row.get("biaya_ops"):
                    bp = str(row["biaya_ops"]).strip()
                    if bp:
                        biaya_parts.append(f"ops:{bp}")
                if row.get("biaya_ops_ket"):
                    bpk = str(row["biaya_ops_ket"]).strip()
                    if bpk:
                        biaya_parts.append(f"keterangan:{bpk}")
                if row.get("biaya_part_service"):
                    bps = str(row["biaya_part_service"]).strip()
                    if bps:
                        biaya_parts.append(f"part_service:{bps}")
            except Exception as e:
                res["errors"].append(f"Baris {row['_row']}: {e}")
                continue
            existing = self.db.query(Mobil).filter(Mobil.nomor_plat == plat).first()
            if dry:
                res["updated" if existing else "created"] += 1
                continue
            if existing:
                existing.merek = merek
                existing.model = model
                existing.tahun = tahun
                existing.warna = warna
                existing.harga_beli = hb
                if hj is not None:
                    existing.harga_jual = hj
                existing.status = st
                existing.tipe_kepemilikan = own
                existing.nama_investor = self._cell(row.get("nama_investor"))
                existing.nomor_mesin = no_mesin
                existing.nomor_rangka = no_rangka
                if ni > 0:
                    existing.nominal_investor = ni
                if pi > 0:
                    existing.persentase_investor = pi
                existing.transmisi = self._cell(row.get("transmisi"))
                existing.bahan_bakar = self._cell(row.get("bahan_bakar"))
                if row.get("kilometer") not in (None, ""):
                    existing.kilometer = self._int(row.get("kilometer"), 0)
                # Append biaya info to catatan
                base_catatan = self._cell(row.get("catatan")) or ""
                if biaya_parts:
                    base_catatan = (base_catatan + " | " if base_catatan else "") + "; ".join(biaya_parts)
                existing.catatan = base_catatan if base_catatan else None
                res["updated"] += 1
            else:
                if not kode:
                    kode = self._next_kode("MBL", Mobil)
                token = uuid.uuid4().hex
                base_catatan = self._cell(row.get("catatan")) or ""
                if biaya_parts:
                    base_catatan = (base_catatan + " | " if base_catatan else "") + "; ".join(biaya_parts)
                self.db.add(
                    Mobil(
                        kode=kode,
                        public_gallery_token=token,
                        merek=merek,
                        model=model,
                        tahun=tahun,
                        warna=warna,
                        nomor_plat=plat,
                        harga_beli=hb,
                        harga_jual=hj,
                        status=st,
                        tanggal_masuk=tgl,
                        tipe_kepemilikan=own,
                        nama_investor=self._cell(row.get("nama_investor")),
                        nomor_mesin=no_mesin,
                        nomor_rangka=no_rangka,
                        nominal_investor=ni,
                        persentase_investor=pi,
                        transmisi=self._cell(row.get("transmisi")),
                        bahan_bakar=self._cell(row.get("bahan_bakar")),
                        kilometer=self._int(row.get("kilometer"), 0) if row.get("kilometer") not in (None, "") else None,
                        status_bayar_beli=PaymentStatus.LUNAS,
                        metode_bayar_beli=PaymentMethod.TUNAI,
                        catatan=base_catatan if base_catatan else None,
                        created_by=user_id,
                    )
                )
                res["created"] += 1
        return res

    def _run(self, file_content: bytes, dry: bool, user_id: Optional[int], batch_id: Optional[str] = None) -> Dict[str, Any]:
        batch_id = batch_id or uuid.uuid4().hex[:8]
        parsed, unknown_sheets = self._parse_workbook(file_content)
        results: Dict[str, Any] = {
            "batch_id": batch_id,
            "dry_run": dry,
            "sheets": {},
            "ok": True,
            "unknown_sheets": unknown_sheets,
        }
        apply_map = {
            "customers": lambda rows: self._apply_customers(rows, dry),
            "suppliers": lambda rows: self._apply_suppliers(rows, dry),
            "spare_parts": lambda rows: self._apply_spare_parts(rows, dry),
            "jasa_servis": lambda rows: self._apply_jasa_servis(rows, dry),
            "karyawan": lambda rows: self._apply_karyawan(rows, dry),
            "kas_opening": lambda rows: self._apply_kas_opening(rows, dry, batch_id, user_id),
            "asset": lambda rows: self._apply_aset(rows, dry),
            "hutang_opening": lambda rows: self._apply_hutang_opening(rows, dry, batch_id, user_id),
            "piutang_opening": lambda rows: self._apply_piutang_opening(rows, dry, batch_id, user_id),
            "armada": lambda rows: self._apply_armada(rows, dry),
            "supir": lambda rows: self._apply_supir(rows, dry),
            "mobil": lambda rows: self._apply_mobil(rows, dry, user_id),
        }
        try:
            for sheet in SHEET_ORDER:
                if sheet == "_INSTRUKSI" or sheet not in parsed:
                    continue
                rows = parsed[sheet]
                if sheet == "neraca_check":
                    # Read-only: store parsed rows for later verification
                    # Row 1 = headers, Row 2 = Excel formulas (auto-sum, skipped in parse)
                    # Row 3+ = user manual input
                    results["sheets"][sheet] = {"rows": len(rows), "data": rows, "errors": []}
                    continue
                sheet_res = apply_map[sheet](rows)
                sheet_res["rows"] = len(rows)
                results["sheets"][sheet] = sheet_res
                if sheet_res.get("errors"):
                    results["ok"] = False
            if dry:
                self.db.rollback()
            else:
                if not results["ok"]:
                    self.db.rollback()
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "message": "Import dibatalkan karena ada error validasi",
                            "batch_id": batch_id,
                            "sheets": results["sheets"],
                        },
                    )
                self.db.commit()
        except HTTPException:
            self.db.rollback()
            raise
        except Exception as e:
            self.db.rollback()
            raise HTTPException(status_code=400, detail=f"Import gagal: {e}")

        # Post-import neraca verification
        results["neraca_verification"] = self._verify_neraca(parsed, results)
        return results

    def _verify_neraca(self, parsed: Dict[str, List[Dict[str, Any]]], results: Dict[str, Any]) -> Dict[str, Any]:
        """Cross-check imported data against neraca_check sheet expectations.
        
        Auto-sums values from other sheets (kas_opening, piutang_opening, hutang_opening, asset, mobil)
        and compares against user-provided expected values in neraca_check sheet.
        Matches the logic from frontend neraca.tsx for verification.
        """
        warnings: List[str] = []
        threshold = Decimal("100")

        def _sum_all_rows(sheet_name: str, field: str) -> Decimal:
            sheet_data = results["sheets"].get(sheet_name, {})
            rows = sheet_data.get("data", [])
            if not rows:
                rows = parsed.get(sheet_name, [])
            total = Decimal("0")
            for row in rows:
                try:
                    total += self._dec(row.get(field, "0"))
                except Exception:
                    continue
            return total

        # Auto-sum from imported sheets (same as neraca.tsx logic)
        total_kas = _sum_all_rows("kas_opening", "nominal")
        total_piutang = _sum_all_rows("piutang_opening", "nominal")
        total_hutang = _sum_all_rows("hutang_opening", "nominal")
        total_aset = _sum_all_rows("asset", "harga_beli")
        total_mobil = _sum_all_rows("mobil", "harga_beli")
        # Persediaan sparepart = sum(stok x harga_beli) per row (matches base.py part_stock)
        total_sparepart = Decimal("0")
        for row in parsed.get("spare_parts", []):
            try:
                total_sparepart += self._dec(row.get("stok", "0")) * self._dec(row.get("harga_beli", "0"))
            except Exception:
                continue

        # Break down kas by jenis_kas column (matching neraca.tsx breakdown)
        kas_tunai = Decimal("0")
        kas_bank = Decimal("0")
        kas_unit_bengkel = Decimal("0")
        kas_unit_ja = Decimal("0")
        kas_unit_mobil = Decimal("0")
        
        kas_rows = parsed.get("kas_opening", [])
        for row in kas_rows:
            try:
                jenis = str(row.get("jenis_kas", "")).strip().upper()
                nominal = self._dec(row.get("nominal", "0"))
                if jenis in ("KAS_UTAMA", "CASH"):
                    kas_tunai += nominal
                elif jenis == "BANK_UTAMA":
                    kas_bank += nominal
                elif jenis == "KAS_UNIT_BENGKEL":
                    kas_unit_bengkel += nominal
                elif jenis == "KAS_UNIT_JASA_ANGKUT":
                    kas_unit_ja += nominal
                elif jenis == "KAS_UNIT_MOBIL":
                    kas_unit_mobil += nominal
            except Exception:
                continue

        total_kas_computed = kas_tunai + kas_bank + kas_unit_bengkel + kas_unit_ja + kas_unit_mobil

        # Break down piutang by jenis_piutang (matching neraca.tsx breakdown)
        piutang_karyawan = Decimal("0")
        piutang_usaha = Decimal("0")
        piutang_mobil = Decimal("0")
        piutang_ja = Decimal("0")
        piutang_lainnya = Decimal("0")
        
        piutang_rows = parsed.get("piutang_opening", [])
        for row in piutang_rows:
            try:
                jenis = str(row.get("unit", "")).strip().upper()
                nominal = self._dec(row.get("nominal", "0"))
                if jenis == "KASBON":
                    piutang_karyawan += nominal
                elif jenis == "BENGKEL":
                    piutang_usaha += nominal
                elif jenis == "JUAL_BELI_MOBIL":
                    piutang_mobil += nominal
                elif jenis == "JASA_ANGKUT":
                    piutang_ja += nominal
                else:
                    piutang_lainnya += nominal
            except Exception:
                continue

        total_piutang_computed = piutang_karyawan + piutang_usaha + piutang_mobil + piutang_ja + piutang_lainnya

        # Break down hutang by jenis_hutang (matching neraca.tsx breakdown)
        hutang_part = Decimal("0")
        hutang_mobil = Decimal("0")
        hutang_investor = Decimal("0")
        hutang_lainnya = Decimal("0")
        hutang_ja = Decimal("0")
        
        hutang_rows = parsed.get("hutang_opening", [])
        for row in hutang_rows:
            try:
                jenis = str(row.get("unit", "")).strip().upper()
                nominal = self._dec(row.get("nominal", "0"))
                if jenis == "BENGKEL":
                    hutang_part += nominal
                elif jenis == "JUAL_BELI_MOBIL":
                    hutang_mobil += nominal
                elif jenis == "INVESTOR":
                    hutang_investor += nominal
                elif jenis == "JASA_ANGKUT":
                    hutang_ja += nominal
                else:
                    hutang_lainnya += nominal
            except Exception:
                continue

        total_hutang_computed = hutang_part + hutang_mobil + hutang_investor + hutang_lainnya + hutang_ja

        # Get expected values from neraca_check sheet
        neraca_rows = parsed.get("neraca_check", [])
        expected: Dict[str, Optional[Decimal]] = {}
        for row in neraca_rows:
            expected["kas_tunai"] = self._dec(row.get("kas_tunai"), "0")
            expected["kas_bank"] = self._dec(row.get("kas_bank"), "0")
            expected["unit_cash"] = self._dec(row.get("unit_cash"), "0")
            expected["total_kas_bank"] = self._dec(row.get("total_kas_bank"), "0")
            expected["piutang_lainnya"] = self._dec(row.get("piutang_lainnya"), "0")
            expected["piutang_karyawan"] = self._dec(row.get("piutang_karyawan"), "0")
            expected["piutang_usaha"] = self._dec(row.get("piutang_usaha"), "0")
            expected["piutang_mobil"] = self._dec(row.get("piutang_mobil"), "0")
            expected["piutang_jasa_angkut"] = self._dec(row.get("piutang_jasa_angkut"), "0")
            expected["total_piutang"] = self._dec(row.get("total_piutang"), "0")
            expected["persediaan_sparepart"] = self._dec(row.get("persediaan_sparepart"), "0")
            expected["stok_mobil"] = self._dec(row.get("stok_mobil"), "0")
            expected["stok_mobil_detail_harga_beli"] = self._dec(row.get("stok_mobil_detail_harga_beli"), "0")
            expected["stok_mobil_detail_biaya_persiapan"] = self._dec(row.get("stok_mobil_detail_biaya_persiapan"), "0")
            expected["stok_mobil_detail_perbaikan_external"] = self._dec(row.get("stok_mobil_detail_perbaikan_external"), "0")
            expected["stok_mobil_detail_perbaikan_internal"] = self._dec(row.get("stok_mobil_detail_perbaikan_internal"), "0")
            expected["total_aset_tetap"] = self._dec(row.get("total_aset_tetap"), "0")
            expected["setoran_modal"] = self._dec(row.get("setoran_modal"), "0")
            expected["setoran_modal_kas"] = self._dec(row.get("setoran_modal_kas"), "0")
            expected["modal_non_kas"] = self._dec(row.get("modal_non_kas"), "0")
            expected["modal_persediaan"] = self._dec(row.get("modal_persediaan"), "0")
            expected["modal_stok_mobil"] = self._dec(row.get("modal_stok_mobil"), "0")
            expected["modal_aset_tetap"] = self._dec(row.get("modal_aset_tetap"), "0")
            expected["laba_ditahan"] = self._dec(row.get("laba_ditahan"), "0")
            expected["prive"] = self._dec(row.get("prive"), "0")
            expected["total_hutang"] = self._dec(row.get("total_hutang"), "0")
            expected["hutang_part"] = self._dec(row.get("hutang_part"), "0")
            expected["hutang_mobil"] = self._dec(row.get("hutang_mobil"), "0")
            expected["hutang_investor"] = self._dec(row.get("hutang_investor"), "0")
            expected["hutang_lainnya"] = self._dec(row.get("hutang_lainnya"), "0")
            expected["hutang_jasa_angkut"] = self._dec(row.get("hutang_jasa_angkut"), "0")

        # Recomputed values (auto-sum from imported sheets)
        computed_total_aktiva = total_kas_computed + total_piutang_computed + total_sparepart + total_aset + total_mobil

        computed: Dict[str, Any] = {
            "kas_tunai": float(kas_tunai),
            "kas_bank": float(kas_bank),
            "unit_cash": float(kas_unit_bengkel + kas_unit_ja + kas_unit_mobil),
            "total_kas": float(total_kas_computed),
            "piutang_karyawan": float(piutang_karyawan),
            "piutang_usaha": float(piutang_usaha),
            "piutang_mobil": float(piutang_mobil),
            "piutang_jasa_angkut": float(piutang_ja),
            "piutang_lainnya": float(piutang_lainnya),
            "total_piutang": float(total_piutang_computed),
            "persediaan_sparepart": float(total_sparepart),
            "total_aset_tetap": float(total_aset),
            "total_mobil": float(total_mobil),
            "hutang_part": float(hutang_part),
            "hutang_mobil": float(hutang_mobil),
            "hutang_investor": float(hutang_investor),
            "hutang_lainnya": float(hutang_lainnya),
            "hutang_jasa_angkut": float(hutang_ja),
            "total_hutang": float(total_hutang_computed),
            "total_aktiva": float(computed_total_aktiva),
        }

        # Calculate expected totals if user provided component values
        if expected.get("total_aset_tetap") and expected["total_aset_tetap"] > 0:
            expected_total_aktiva = (
                expected.get("total_kas_bank", Decimal("0"))
                + expected.get("total_piutang", Decimal("0"))
                + expected.get("persediaan_sparepart", Decimal("0"))
                + expected.get("stok_mobil", Decimal("0"))
                + expected["total_aset_tetap"]
            )
        else:
            expected_total_aktiva = expected.get("total_aktiva", Decimal("0"))

        # Konsep saldo awal: Modal = Total Aktiva - Total Hutang (modal sebagai plug)
        expected_total_hutang_exp = expected.get("total_hutang", Decimal("0"))
        if expected_total_aktiva > 0 and expected_total_hutang_exp > 0:
            expected_total_modal = expected_total_aktiva - expected_total_hutang_exp
        else:
            expected_total_modal = expected.get("setoran_modal", Decimal("0"))
            if expected_total_modal == 0:
                expected_total_modal = (
                    expected.get("setoran_modal_kas", Decimal("0"))
                    + expected.get("modal_non_kas", Decimal("0"))
                    + expected.get("laba_ditahan", Decimal("0"))
                    - expected.get("prive", Decimal("0"))
                )
        expected_total_pasiva = expected_total_hutang_exp + expected_total_modal

        # Compare each expected value against computed
        check_fields = [
            ("total_kas", expected.get("total_kas_bank"), computed["total_kas"]),
            ("total_piutang", expected.get("total_piutang"), computed["total_piutang"]),
            ("total_hutang", expected.get("total_hutang"), computed["total_hutang"]),
            ("total_aset_tetap", expected.get("total_aset_tetap"), computed["total_aset_tetap"]),
            ("persediaan_sparepart", expected.get("persediaan_sparepart"), computed["persediaan_sparepart"]),
            ("stok_mobil", expected.get("stok_mobil"), computed["total_mobil"]),
        ]

        for label, exp_val, comp_val in check_fields:
            if exp_val and exp_val > 0:
                diff = abs(exp_val - Decimal(str(comp_val)))
                if diff > threshold:
                    warnings.append(
                        f"{label}: auto-sum {comp_val:,.0f} vs expected {exp_val:,.0f} (selisih {diff:,.0f})"
                    )

        # Balance check: compare total_aktiva vs total_pasiva
        if expected_total_aktiva and expected_total_aktiva > 0:
            selisih = computed_total_aktiva - expected_total_aktiva
            if abs(selisih) > threshold:
                warnings.append(
                    f"Total Aktiva: computed {computed_total_aktiva:,.0f} vs expected {expected_total_aktiva:,.0f} (selisih {selisih:,.0f})"
                )

        if expected_total_pasiva and expected_total_pasiva > 0:
            selisih = abs(computed_total_aktiva - expected_total_pasiva)
            if selisih > threshold:
                warnings.append(
                    f"Neraca tidak seimbang: Aktiva {computed_total_aktiva:,.0f} vs Pasiva {expected_total_pasiva:,.0f} (selisih {selisih:,.0f})"
                )

        return {
            "computed": computed,
            "expected": {k: float(v) if v else None for k, v in expected.items()},
            "is_balanced": len(warnings) == 0,
            "warnings": warnings,
        }

    def preview(self, file_content: bytes, user_id: Optional[int] = None) -> Dict[str, Any]:
        return self._run(file_content, dry=True, user_id=user_id)

    def commit(self, file_content: bytes, user_id: Optional[int] = None, batch_id: Optional[str] = None) -> Dict[str, Any]:
        return self._run(file_content, dry=False, user_id=user_id, batch_id=batch_id)
